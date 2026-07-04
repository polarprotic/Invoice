"""
Invoice Extraction Pipeline -- Local OCR + Qwen (Ollama), with dynamic fields
------------------------------------------------------------------------------
Same job as before, but no longer forces every document into one rigid
column layout. Real invoices/cheques/vouchers vary a lot (discount, GST,
freight, PO number, due date, TDS, etc.), so:

  - CORE fields (document_type, date, payer, payee, amount, reference_no,
    bank, account_no, notes) are always extracted and always appear as the
    first columns -- these exist on almost every document type here.
  - Anything else the model finds on a document (discount, tax, GST number,
    due date, PO number, freight, TDS, cheque validity date, etc.) goes into
    an "extra_fields" object. Per OUTPUT FILE, all extra keys seen across
    that file's records are unioned into extra columns appended after the
    core ones -- so a file with 3 cheques and 1 GST invoice gets one sheet
    with "Discount"/"GST No" columns filled only where relevant, blank
    elsewhere, instead of losing that data or forcing a one-size-fits-all
    schema across your whole batch.

REQUIREMENTS:
  - Ollama running locally or on your LAN, with the model already pulled:
        ollama pull qwen3-vl:8b
  - Tesseract OCR installed (Windows: UB-Mannheim installer). If it's not
    on your PATH, set TESSERACT_CMD below to the .exe path.
  - pip install pytesseract pymupdf pillow pandas openpyxl requests

USAGE:
    python invoice_pipeline_ollama.py

    (reads from ./input_img, writes to ./invoice_output/<filename>/)
"""

import os
import io
import re
import json
import time
import glob
import base64
import shutil

import requests
import pytesseract
from PIL import Image
import fitz  # PyMuPDF
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- CONFIG ----------
INPUT_DIR = "input_img"
OUTPUT_ROOT = "invoice_output"

OLLAMA_HOST = "http://localhost:11434"
OLLAMA_MODEL = "qwen3-vl:8b"

NUM_CTX = 8192

TESSERACT_CMD = None
if TESSERACT_CMD:
    pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD

SLEEP_BETWEEN_CALLS = 1
MAX_IMAGE_DIMENSION = 1600
MAX_RETRIES = 3
REQUEST_TIMEOUT = 420  # seconds

# Cap on how many distinct extra-field columns a single file's sheet can
# grow to. Guards against a garbled model response spraying dozens of junk
# keys into the header row. Anything beyond the cap collapses into "notes".
# Raised from 12 -- the exhaustive prompt below routinely surfaces 8-15
# genuinely distinct fields (time, currency, subtotal, tax, payment_method,
# rep, addresses, etc.) even on a single simple receipt.
MAX_EXTRA_COLUMNS = 25

# Keys (after normalize_key) that should be written as real numbers rather
# than text, so Excel SUM()/number formatting works on them instead of
# silently treating them as blank.
NUMERIC_KEYS = {"subtotal", "tax", "discount", "tds", "freight",
                 "balance_remaining"}
# -----------------------------

CORE_COLUMNS = ["source_file", "document_type", "date", "payer", "payee",
                "amount", "reference_no", "bank", "account_no", "notes"]

# Known synonyms get normalized to one canonical key so "Discount %" and
# "discount_percent" don't become two separate columns. Match is
# case/whitespace-insensitive against the lowercased, underscored key the
# model returns.
FIELD_ALIASES = {
    "discount": "discount",
    "discount_amount": "discount",
    "discount_percent": "discount",
    "gst": "gst_no",
    "gst_no": "gst_no",
    "gst_number": "gst_no",
    "gstin": "gst_no",
    "tax": "tax",
    "tax_amount": "tax",
    "tds": "tds",
    "due_date": "due_date",
    "payment_due_date": "due_date",
    "po_number": "po_number",
    "purchase_order_no": "po_number",
    "freight": "freight",
    "shipping_charges": "freight",
    "cheque_validity_date": "cheque_validity_date",
    "ifsc_code": "ifsc_code",
    "ifsc": "ifsc_code",
}


def normalize_key(k):
    k = re.sub(r"[^a-z0-9]+", "_", k.strip().lower()).strip("_")
    return FIELD_ALIASES.get(k, k)


BASE_PROMPT = """Extract structured data from this Indian bank/advertising payment
document. It may contain a MIX of typed/printed text and handwritten text in
the same image (e.g. a printed form filled in by hand, or a receipt with a
handwritten section next to a printed letterhead). Read both carefully,
using the actual image as ground truth.

Return a JSON array. Each element = ONE distinct financial transaction or
record found anywhere in the image (a single image may show multiple cheques,
receipts, or line-items -- list each one separately, do not merge them).

For each record extract these CORE fields (always include all of them, use ""
if missing/illegible):
- document_type: e.g. "Cheque", "Deposit Slip", "NEFT Advice", "Payment Voucher",
  "Release Order", "Receipt", "Tax Invoice", "Invoice"
- date: exactly as written (date only -- put any separate time value in extra_fields, see below)
- payer: who is paying / sending money / the client being billed
- payee: who receives the money / the business issuing the document
- amount: the FINAL/net payable or total amount, number only, no currency
  symbols, no commas, no units (e.g. "540" not "USD540" or "540.00 USD")
- reference_no: cheque no. / UTR / transaction ID / receipt no. / invoice no. / bill no. --
  whichever is present (combine if multiple, comma separated)
- bank: bank name involved (leave "" if not a banking document)
- account_no: as shown, including masked digits if partially hidden
- notes: only for genuinely unstructured remarks that don't fit as a labeled
  field anywhere else (e.g. a free-text comment) -- do NOT use notes as a
  dumping ground for labeled fields, those belong in extra_fields below

CRITICAL -- BE EXHAUSTIVE: this document may be one of thousands of
different formats (invoices, receipts, cheques, vouchers, advices) and each
one can carry different labeled information. Do not limit yourself to a
fixed list. Scan the ENTIRE document top to bottom and capture EVERY
distinct labeled piece of information you can see as a key in
"extra_fields", even if it seems minor or you're unsure it matters.
Concretely, always check for (include only if actually present on the
document):
- time (if a timestamp with a time component appears, separate from date)
- currency / currency_unit (e.g. "USD", "INR") if the amount has one
- subtotal, tax, discount, gst_no, tds, freight, balance_remaining,
  amount_in_words
- invoice_no (if distinct from reference_no), po_number, due_date,
  payment_method, rep / prepared_by / issued_by, client_address,
  company_address, company_phone, company_email, ifsc_code,
  branch, cheque_validity_date
- ANY other labeled field visible on the document that isn't covered above
  -- invent a short snake_case key for it rather than skipping it or
  folding it into notes.

Use short snake_case keys and plain values -- strip currency symbols/commas
from numeric-looking values (put the currency itself in a separate
"currency" key if shown). Only include keys for data actually present in
THIS document -- do not invent data, and leave "extra_fields" as {} only if
there is truly nothing beyond the core fields.

Return ONLY the JSON array -- no markdown fences, no explanation, no
commentary before or after the array. Example shape:
[{"document_type": "...", "date": "...", "payer": "...", "payee": "...",
  "amount": "540", "reference_no": "...", "bank": "...", "account_no": "...",
  "notes": "...", "extra_fields": {"time": "11:10", "currency": "USD",
  "subtotal": "540", "tax": "0", "balance_remaining": "0",
  "payment_method": "Bank Transfer", "rep": "J. Nolan"}}]"""


def load_pages(path):
    ext = path.lower().rsplit(".", 1)[-1]
    if ext == "pdf":
        doc = fitz.open(path)
        return [Image.open(io.BytesIO(page.get_pixmap(dpi=200).tobytes("png"))) for page in doc]
    return [Image.open(path).convert("RGB")]


def ocr_hint_text(images):
    chunks = []
    for i, img in enumerate(images):
        try:
            text = pytesseract.image_to_string(img, config="--oem 3 --psm 6").strip()
        except Exception as e:
            text = f"[OCR unavailable: {e}]"
        chunks.append(f"--- OCR of page {i + 1} (typed text only, may contain errors,\n"
                       f"    handwriting likely garbled -- use as a hint, not ground truth) ---\n{text}")
    return "\n\n".join(chunks)


def build_prompt(ocr_hint):
    return (f"{BASE_PROMPT}\n\n"
            f"For reference, here is raw OCR text pulled from this document. It's "
            f"machine-generated and may have errors -- especially anything "
            f"handwritten -- so treat it only as a hint for reading dense typed "
            f"sections, not as the final answer:\n\n{ocr_hint}\n\n"
            f"Now return ONLY the JSON array as instructed above.")


def resize_for_model(img, max_dim=MAX_IMAGE_DIMENSION):
    w, h = img.size
    longest = max(w, h)
    if longest <= max_dim:
        return img
    scale = max_dim / longest
    new_size = (int(w * scale), int(h * scale))
    return img.resize(new_size, Image.LANCZOS)


def pil_to_b64(img):
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return base64.b64encode(buf.getvalue()).decode("utf-8")


def extract_json(text):
    text = text.strip()

    # qwen3-vl (and other reasoning-tuned models) sometimes emit a
    # <think>...</think> reasoning block before the actual answer. Strip it
    # so it doesn't get parsed as (or block detection of) the JSON array.
    text = re.sub(r"<think>.*?</think>", "", text, flags=re.DOTALL).strip()

    try:
        return json.loads(text)
    except Exception:
        pass

    fenced = re.search(r"```(?:json)?\s*(.*?)```", text, re.DOTALL)
    if fenced:
        try:
            return json.loads(fenced.group(1).strip())
        except Exception:
            pass

    start = text.find("[")
    if start != -1:
        depth = 0
        for i in range(start, len(text)):
            if text[i] == "[":
                depth += 1
            elif text[i] == "]":
                depth -= 1
                if depth == 0:
                    candidate = text[start:i + 1]
                    try:
                        return json.loads(candidate)
                    except Exception:
                        break

    return None


def check_ollama():
    try:
        resp = requests.get(f"{OLLAMA_HOST}/api/tags", timeout=10)
        resp.raise_for_status()
        models = [m["name"] for m in resp.json().get("models", [])]
    except requests.exceptions.ConnectionError:
        print(f"ERROR: can't reach Ollama at {OLLAMA_HOST}")
        print("  - Is Ollama running there? Try: ollama serve")
        print(f"  - Is {OLLAMA_HOST} the right host/port for this machine?")
        return False
    except Exception as e:
        print(f"ERROR: unexpected response checking Ollama: {e}")
        return False

    if not any(OLLAMA_MODEL in m for m in models):
        print(f"ERROR: model '{OLLAMA_MODEL}' not found on {OLLAMA_HOST}.")
        print(f"  Pull it with: ollama pull {OLLAMA_MODEL}")
        print(f"  Currently available: {', '.join(models) if models else '(none)'}")
        return False

    return True


def extract(path, images, debug_dir=None):
    ocr_hint = ocr_hint_text(images)
    prompt = build_prompt(ocr_hint)
    images_b64 = [pil_to_b64(resize_for_model(img)) for img in images]

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = requests.post(
                f"{OLLAMA_HOST}/api/generate",
                json={
                    "model": OLLAMA_MODEL,
                    "prompt": prompt,
                    "images": images_b64,
                    "stream": False,
                    "options": {"num_ctx": NUM_CTX, "temperature": 0.1},
                },
                timeout=REQUEST_TIMEOUT,
            )
            resp.raise_for_status()
            raw = resp.json().get("response", "")
            parsed = extract_json(raw)
            if parsed is not None:
                return parsed
            print(f"    [parse failed, retry {attempt}/{MAX_RETRIES}] model didn't return valid JSON")
            if debug_dir:
                # Dump what the model actually said -- essential for
                # figuring out *why* parsing failed instead of guessing.
                os.makedirs(debug_dir, exist_ok=True)
                dump_path = os.path.join(debug_dir, f"raw_response_attempt{attempt}.txt")
                with open(dump_path, "w", encoding="utf-8") as fh:
                    fh.write(raw)
                print(f"        raw response saved to {dump_path}")

        except requests.exceptions.ConnectionError:
            print(f"    [connection error, retry {attempt}/{MAX_RETRIES}] "
                  f"is Ollama still running at {OLLAMA_HOST}?")
        except requests.exceptions.Timeout:
            print(f"    [timeout after {REQUEST_TIMEOUT}s, retry {attempt}/{MAX_RETRIES}]")
        except Exception as e:
            print(f"    [retry {attempt}/{MAX_RETRIES}] {e}")

        time.sleep(5 * attempt)

    print(f"    [FAILED] {os.path.basename(path)} -- skipped, needs manual entry")
    return []


def flatten_records(results, source_file):
    """Split each raw model record into its core dict + normalized
    extra_fields dict. No capping here -- this file's fields get unioned
    with every other file's fields first, and the cap (if any) is applied
    once globally in main() so column layout stays consistent batch-wide."""
    core_rows = []

    for r in results:
        core = {col: r.get(col, "") for col in CORE_COLUMNS}
        core["source_file"] = source_file

        raw_extra = r.get("extra_fields") or {}
        norm_extra = {}
        if isinstance(raw_extra, dict):
            for k, v in raw_extra.items():
                nk = normalize_key(str(k))
                if nk:
                    norm_extra[nk] = v

        core["_extra"] = norm_extra
        core_rows.append(core)

    return core_rows


def build_global_extra_columns(all_core_rows):
    """Union extra_fields keys across EVERY row in the whole batch, in
    first-seen order. Cap applied once, globally -- fields beyond the cap
    fold into that row's notes instead of getting a column, so a rare
    one-off field on invoice #847 doesn't blow up every sheet's width, but
    common fields shared across many invoices still get their own column
    no matter how far into the batch they first appear."""
    freq = {}
    order = []
    for core in all_core_rows:
        for k in core.get("_extra", {}):
            if k not in freq:
                freq[k] = 0
                order.append(k)
            freq[k] += 1

    # Prioritize fields that appear across more documents (more likely to
    # be a genuine recurring field, not a one-off OCR/model quirk), tie-broken
    # by first-seen order.
    ranked = sorted(order, key=lambda k: (-freq[k], order.index(k)))
    kept = ranked[:MAX_EXTRA_COLUMNS]
    dropped = set(ranked[MAX_EXTRA_COLUMNS:])

    if dropped:
        for core in all_core_rows:
            extra = core.get("_extra", {})
            overflow = {k: v for k, v in extra.items() if k in dropped}
            if overflow:
                extra_str = "; ".join(f"{k}: {v}" for k, v in overflow.items())
                core["notes"] = (core["notes"] + "; " if core["notes"] else "") + extra_str

    # Return kept columns in original first-seen order (nicer to read than
    # frequency order in the sheet itself).
    return [k for k in order if k in set(kept)]


def style_workbook(path, amt_col_letter, n_core_cols, n_extra_cols):
    wb = load_workbook(path)
    ws = wb.active
    n_cols = ws.max_column

    header_fill = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
    extra_header_fill = PatternFill("solid", start_color="2E7D32", end_color="2E7D32")
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, n_cols + 1):
        c = ws.cell(row=1, column=col)
        # Extra (dynamic) columns get a different header color so it's
        # visually obvious which fields are core vs. document-specific.
        c.fill = extra_header_fill if col > n_core_cols else header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    for r in range(2, ws.max_row + 1):
        for col in range(1, n_cols + 1):
            c = ws.cell(row=r, column=col)
            c.font = Font(name="Arial", size=10)
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if col == amt_col_letter:
                c.number_format = "#,##0.00"
                c.alignment = Alignment(horizontal="right", vertical="center")

    core_widths = [30, 22, 14, 26, 26, 14, 36, 20, 18, 36]
    for i, w in enumerate(core_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for i in range(n_core_cols + 1, n_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18
    ws.freeze_panes = "A2"

    if ws.max_row > 1:
        last_row = ws.max_row
        total_row = last_row + 2
        col_letter = get_column_letter(amt_col_letter)
        # Label goes one column to the left of amount (was previously
        # hardcoded to column 5, which only matched the old fixed schema
        # and could misalign or overlap the amount column itself).
        label_col = max(1, amt_col_letter - 1)
        ws.cell(row=total_row, column=label_col, value="TOTAL").font = Font(name="Arial", bold=True)
        tc = ws.cell(row=total_row, column=amt_col_letter,
                     value=f"=SUM({col_letter}2:{col_letter}{last_row})")
        tc.number_format = "#,##0.00"
        tc.font = Font(name="Arial", bold=True)

    wb.save(path)


def write_master_excel(all_core_rows, extra_columns, out_path):
    """One combined sheet with every record from every file in the batch,
    using the same global column set as the per-file sheets -- this is the
    one to open when you want to filter/pivot/sum across the whole batch."""
    write_excel(all_core_rows, extra_columns, out_path)


def write_excel(core_rows, extra_columns, out_path):
    """Build one sheet: core columns first, then whatever extra fields
    actually showed up across this file's records (union, first-seen
    order). Cells left blank where a given record didn't have that field."""
    all_columns = CORE_COLUMNS + extra_columns
    rows = []
    for core in core_rows:
        row = {col: core.get(col, "") for col in CORE_COLUMNS}
        extra = core.get("_extra", {})
        for ec in extra_columns:
            row[ec] = extra.get(ec, "")
        rows.append(row)

    df = pd.DataFrame(rows, columns=all_columns)

    # Coerce amount + known numeric extra fields from text to real numbers
    # (model returns strings) -- otherwise Excel SUM() silently treats them
    # as blank/zero instead of erroring, which is easy to miss.
    def to_number(v):
        if v in ("", None):
            return None
        try:
            return float(str(v).replace(",", "").strip())
        except ValueError:
            return v  # leave as-is if it genuinely isn't numeric

    df["amount"] = df["amount"].apply(to_number)
    for col in extra_columns:
        if col in NUMERIC_KEYS:
            df[col] = df[col].apply(to_number)

    try:
        df.to_excel(out_path, index=False)
        amt_col_idx = CORE_COLUMNS.index("amount") + 1
        style_workbook(out_path, amt_col_idx, len(CORE_COLUMNS), len(extra_columns))
    except PermissionError:
        raise PermissionError(
            f"'{out_path}' is open in Excel (or similar) -- close it and re-run "
            f"the script later to pick this file up."
        )


def main():
    print("=" * 60)
    print("CONFIG IN USE (check this matches what you expect):")
    print(f"  OLLAMA_HOST         = {OLLAMA_HOST}")
    print(f"  OLLAMA_MODEL        = {OLLAMA_MODEL}")
    print(f"  MAX_IMAGE_DIMENSION = {MAX_IMAGE_DIMENSION}")
    print(f"  REQUEST_TIMEOUT     = {REQUEST_TIMEOUT}s")
    print(f"  NUM_CTX             = {NUM_CTX}")
    print(f"  MAX_EXTRA_COLUMNS   = {MAX_EXTRA_COLUMNS}")
    print("=" * 60 + "\n")

    if not check_ollama():
        return

    files = sorted(
        f for f in glob.glob(os.path.join(INPUT_DIR, "*"))
        if f.lower().endswith((".jpg", ".jpeg", ".png", ".pdf"))
    )
    if not files:
        print(f"No .jpg/.jpeg/.png/.pdf files found in '{INPUT_DIR}/'")
        return

    os.makedirs(OUTPUT_ROOT, exist_ok=True)

    def raw_cache_path(f):
        base_name = os.path.splitext(os.path.basename(f))[0]
        return os.path.join(OUTPUT_ROOT, base_name, "raw_extraction.json")

    def already_extracted(f):
        # Cache the MODEL OUTPUT, not the final xlsx -- the xlsx layout gets
        # rebuilt from cache every run to pick up new columns from newer
        # files, but re-calling the (slow) model on unchanged files is
        # wasted work.
        return os.path.exists(raw_cache_path(f))

    pending = [f for f in files if not already_extracted(f)]
    cached_count = len(files) - len(pending)
    if cached_count:
        print(f"{cached_count} file(s) already extracted (using cached results).")
    if pending:
        print(f"{len(pending)} file(s) to extract with {OLLAMA_MODEL} at {OLLAMA_HOST}.\n")

    failed = []

    # ---- PASS 1: extract (or load from cache) every file's raw records ----
    for i, f in enumerate(pending, 1):
        base_name = os.path.splitext(os.path.basename(f))[0]
        print(f"[{i}/{len(pending)}] Extracting {os.path.basename(f)}")

        folder = os.path.join(OUTPUT_ROOT, base_name)
        os.makedirs(folder, exist_ok=True)

        try:
            shutil.copy2(f, os.path.join(folder, os.path.basename(f)))
            images = load_pages(f)
            results = extract(f, images, debug_dir=folder)
            tmp_path = raw_cache_path(f) + ".tmp"
            with open(tmp_path, "w", encoding="utf-8") as fh:
                json.dump(results, fh, ensure_ascii=False, indent=2)
            os.replace(tmp_path, raw_cache_path(f))  # atomic -- avoids a
            # half-written cache file if the script is killed mid-write
            print(f"    -> {len(results)} record(s) extracted")
        except Exception as e:
            print(f"    [SKIPPED - error] {os.path.basename(f)}: {e}")
            failed.append(os.path.basename(f))

        time.sleep(SLEEP_BETWEEN_CALLS)

    # ---- PASS 2: rebuild every sheet from cache with a consistent, ----
    # ---- batch-wide column layout (covers cached files from earlier runs too) ----
    print("\nBuilding spreadsheets with a consistent column layout across the batch...")

    per_file_rows = {}  # base_name -> core_rows (with _extra intact)
    for f in files:
        cache_path = raw_cache_path(f)
        if not os.path.exists(cache_path):
            continue  # this file failed extraction and has no cache yet
        base_name = os.path.splitext(os.path.basename(f))[0]
        try:
            with open(cache_path, "r", encoding="utf-8") as fh:
                results = json.load(fh)
        except json.JSONDecodeError:
            # Corrupted/truncated cache (e.g. script was killed mid-write on
            # a previous run). Don't crash the whole batch over one bad
            # file -- delete the stale cache so it's re-extracted on the
            # next run, and skip it in this one.
            print(f"    [BAD CACHE] {base_name}: raw_extraction.json is corrupted, "
                  f"deleting it -- re-run the script to re-extract this file.")
            os.remove(cache_path)
            failed.append(os.path.basename(f))
            continue
        per_file_rows[base_name] = flatten_records(results, os.path.basename(f))

    all_core_rows = [row for rows in per_file_rows.values() for row in rows]
    if not all_core_rows:
        print("No successfully extracted records to write.")
    else:
        extra_columns = build_global_extra_columns(all_core_rows)
        print(f"Global extra columns ({len(extra_columns)}): {', '.join(extra_columns) if extra_columns else '(none)'}")

        for base_name, rows in per_file_rows.items():
            out_xlsx = os.path.join(OUTPUT_ROOT, base_name, "Extracted_Invoice.xlsx")
            try:
                write_excel(rows, extra_columns, out_xlsx)
            except PermissionError as e:
                print(f"    [SKIPPED WRITE] {base_name}: {e}")

        master_path = os.path.join(OUTPUT_ROOT, "All_Invoices_Combined.xlsx")
        try:
            write_master_excel(all_core_rows, extra_columns, master_path)
            print(f"\nMaster combined workbook: {master_path} ({len(all_core_rows)} record(s) across {len(per_file_rows)} file(s))")
        except PermissionError as e:
            print(f"    [SKIPPED WRITE] master workbook: {e}")

    print(f"\nDone. Per-file results in '{OUTPUT_ROOT}/<filename>/Extracted_Invoice.xlsx'.")
    if failed:
        print(f"{len(failed)} file(s) failed extraction (re-run the script "
              f"to retry them automatically): {', '.join(failed)}")


if __name__ == "__main__":
    main()