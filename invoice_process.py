"""
Invoice Extraction Pipeline -- Cloud GPT-4o API, CONTINUOUS WATCH-FOLDER
COST-OPTIMIZED VERSION
------------------------------------------------------------------------------
Changes from the original, all aimed at cutting API spend while protecting
handwriting accuracy:

  1. API key now comes from an environment variable, not hardcoded in the
     file (rotate your old key -- it was exposed).
  2. Per-page image "detail" is chosen adaptively:
       - pages that look mostly typed/printed (based on OCR confidence and
         character density) are sent at "low" detail (~4x cheaper per image)
       - pages that look sparse, low-confidence, or likely handwritten keep
         "high" detail, so accuracy on the hard pages is untouched
  3. OCR hint text sent in the prompt is trimmed/filtered: low-confidence
     garbled OCR (mostly noise on handwritten sections) is no longer
     stuffed into the prompt, and hints are capped in length.
  4. MAX_RETRIES reduced 3 -> 2. Retries resend the full image, so they're
     the most expensive kind of waste; illegibility from handwriting is a
     capability issue, not something a 3rd retry usually fixes.
  5. Blank / near-blank PDF pages are detected and skipped before ever
     being sent to the API.
  6. Everything else (prompt content, line-item logic, Excel output,
     watch-folder loop, state file) is unchanged from the original.

OUTPUT: Per-file subfolders in OUTPUT_ROOT.
        For every processed image -> <OUTPUT_ROOT>/<basename>/
            - <basename>.xlsx   (single-file Excel)
            - <basename>.<ext>  (copy of the original input image/PDF)
"""

import os
import io
import re
import json
import time
import base64
import shutil
import signal
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed

import numpy as np
import pytesseract
from PIL import Image
import fitz  # PyMuPDF
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openai import OpenAI

# ---------- CONFIG ----------
OUTPUT_SUFFIX = "_output"
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

OPENAI_MODEL = "gpt-4o-mini"

# SECURITY: never hardcode the key. Set it before running:
#   export OPENAI_API_KEY="sk-..."          (mac/linux)
#   setx OPENAI_API_KEY "sk-..."            (windows)
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "")

API_CONCURRENT_WORKERS = 4
NUM_SAMPLES_PER_FILE = 1
MODEL_TEMPERATURE = 0.2

TESSERACT_CMD = None
if TESSERACT_CMD:
    pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD

MAX_IMAGE_DIMENSION = 2200
MAX_RETRIES = 2                     # was 3 -- retries resend full images, expensive
REQUEST_TIMEOUT = 120
MAX_OUTPUT_TOKENS = 4096
MAX_EXTRA_COLUMNS = 25

NUMERIC_KEYS = {"subtotal", "tax", "discount", "tds", "freight", "balance_remaining"}

# ---- CONTINUOUS CONFIG ----
POLL_INTERVAL_SECONDS = 20
STATE_FILE_NAME = "_pipeline_state.json"

# ---- COST OPTIMIZATION CONFIG ----
# OCR confidence (0-100) above which a page is considered "cleanly typed"
# enough to use low-detail image encoding. Handwriting / noisy scans score
# low here and automatically keep high-detail (full accuracy).
LOW_DETAIL_OCR_CONFIDENCE_THRESHOLD = 75
# Minimum OCR character count for a page to even be considered for low
# detail -- very sparse pages are treated as possibly handwritten/complex.
LOW_DETAIL_MIN_CHAR_COUNT = 120
# OCR hint text is capped to this many characters per page to avoid paying
# for garbled, low-value OCR noise as prompt tokens.
OCR_HINT_MAX_CHARS_PER_PAGE = 1200
# Below this per-page OCR confidence, we don't bother including the OCR
# hint for that page at all (it's mostly noise on handwriting anyway).
OCR_HINT_MIN_CONFIDENCE_TO_INCLUDE = 35
# A PDF page is treated as blank/near-blank (skipped entirely, no API
# cost) if its non-white pixel fraction is below this threshold.
BLANK_PAGE_INK_FRACTION_THRESHOLD = 0.002
# -----------------------------

CORE_COLUMNS = ["source_file", "document_type", "date", "payer", "payee",
                "amount", "reference_no", "bank", "account_no", "notes"]

FIELD_ALIASES = {
    "discount": "discount",
    "discount_amount": "discount",
    "discount_percent": "discount",
    "gst": "gst_no",
    "gst_no": "gst_no",
    "gst_number": "gst_no",
    "gstin": "gst_no",
    "tax": "tax",
    "tax_amount": "tax",
    "tds": "tds",
    "due_date": "due_date",
    "payment_due_date": "due_date",
    "po_number": "po_number",
    "purchase_order_no": "po_number",
    "freight": "freight",
    "shipping_charges": "freight",
    "cheque_validity_date": "cheque_validity_date",
    "ifsc_code": "ifsc_code",
    "ifsc": "ifsc_code",
    "debit_credit": "debit_credit",
    "transaction_description": "transaction_description",
}


def normalize_key(k):
    k = re.sub(r"[^a-z0-9]+", "_", k.strip().lower()).strip("_")
    return FIELD_ALIASES.get(k, k)


BASE_PROMPT = """Extract structured data from this Indian bank/advertising payment
document. It may contain a MIX of typed/printed text and handwritten text in
the same image (e.g. a printed form filled in by hand, or a receipt with a
handwritten section next to a printed letterhead), and it may also be a
SCREENSHOT of a banking portal or payment confirmation page rather than a
physical document -- read screenshots just as carefully as scans. Read
everything carefully, using the actual image as ground truth.

Return a JSON array. Each element = ONE distinct financial transaction or
record found anywhere in the image (a single image may show multiple cheques,
receipts, or line-items -- list each one separately, do not merge them,
UNLESS the document is a single deposit/transaction slip containing an
itemized breakdown table -- see line_items below for that case).

For each record extract these CORE fields (always include all of them, use ""
if missing/illegible):
- document_type: e.g. "Cheque", "Deposit Slip", "NEFT Advice", "Payment Voucher",
  "Release Order", "Receipt", "Tax Invoice", "Invoice", "Bank Portal Screenshot"
- date: exactly as written (date only -- put any separate time value in extra_fields, see below)
- payer: who is paying / sending money / the client being billed
- payee: who receives the money / the business issuing the document
- amount: the FINAL/net payable or total amount, number only, no currency
  symbols, no commas, no units (e.g. "540" not "USD540" or "540.00 USD")
- reference_no: cheque no. / UTR / transaction ID / receipt no. / invoice no. / bill no. / deposit slip no. --
  whichever is present (combine if multiple, comma separated)
- bank: bank name involved (leave "" if not a banking document)
- account_no: as shown, including masked digits if partially hidden
- notes: only for genuinely unstructured remarks that don't fit as a labeled
  field anywhere else (e.g. a free-text comment) -- do NOT use notes as a
  dumping ground for labeled fields, those belong in extra_fields below

LINE ITEMS (for deposit slips / tables with multiple rows under one
transaction, e.g. a bank deposit slip listing many different payee/agent
names each with their own PRT number, place, and amount): put the OVERALL
deposit slip info (deposit slip no, deposit date, client name, total
amount) in the core fields above, and include every individual row as an
object in a "line_items" array in extra_fields, like:
"line_items": [{"name": "...", "place": "...", "prt_no": "...", "prt_date": "...", "amount": "..."}, ...]
Do not skip rows because handwriting is hard to read -- give your best
reading of every row, including ones that are smudged, stamped over, or
partially obscured. If truly illegible, put "illegible" for that field
rather than omitting the row entirely.

CRITICAL -- BE EXHAUSTIVE: this document may be one of thousands of
different formats (invoices, receipts, cheques, vouchers, advices, bank
portal screenshots) and each one can carry different labeled information.
Do not limit yourself to a fixed list. Scan the ENTIRE document/screenshot
top to bottom and capture EVERY distinct labeled piece of information you
can see as a key in "extra_fields", even if it seems minor. Treat every
document with equal scrutiny regardless of how sparse or dense it looks --
a simple-looking document often still has a GST number, IFSC code, due
date, or debit/credit indicator in small print that is easy to skip.
Concretely, always check for (include only if actually present on the
document):
- time (if a timestamp with a time component appears, separate from date)
- currency / currency_unit (e.g. "USD", "INR") if the amount has one
- subtotal, tax, discount, gst_no, tds, freight, balance_remaining,
  amount_in_words
- invoice_no (if distinct from reference_no), po_number, due_date,
  payment_method, rep / prepared_by / issued_by, client_address,
  company_address, company_phone, company_email, ifsc_code,
  branch, cheque_validity_date
- debit_credit (whether the transaction shown is a Debit or a Credit --
  this appears explicitly on many bank portal screenshots/receipts, look
  for a field literally labeled "Debit/Credit" or similar)
- transaction_description (any raw transaction narration/description string
  shown on a bank portal screenshot, e.g. "INF/NEFT/..." style strings --
  copy it in full)
- ANY other labeled field visible on the document that isn't covered above
  -- invent a short snake_case key for it rather than skipping it or
  folding it into notes.

Use short snake_case keys and plain values -- strip currency symbols/commas
from numeric-looking values (put the currency itself in a separate
"currency" key if shown). Only include keys for data actually present in
THIS document -- do not invent data, and leave "extra_fields" as {} only if
there is truly nothing beyond the core fields.

Return ONLY the JSON array -- no markdown fences, no explanation, no
commentary before or after the array. Example shape:
[{"document_type": "...", "date": "...", "payer": "...", "payee": "...",
  "amount": "540", "reference_no": "...", "bank": "...", "account_no": "...",
  "notes": "...", "extra_fields": {"time": "11:10", "currency": "USD",
  "subtotal": "540", "tax": "0", "balance_remaining": "0",
  "payment_method": "Bank Transfer", "rep": "J. Nolan", "debit_credit": "Debit",
  "line_items": [{"name": "ABC", "place": "XYZ", "prt_no": "12345", "prt_date": "28/08", "amount": "3570"}]}}]"""


# -----------------------------------------------------------------------
# IMAGE LOADING / OCR / PROMPT
# -----------------------------------------------------------------------

IMAGE_EXTENSIONS = (".jpg", ".jpeg", ".png", ".pdf")


def discover_input_files(root_dir):
    found = []
    for dirpath, _dirnames, filenames in os.walk(root_dir):
        for fname in filenames:
            if fname.lower().endswith(IMAGE_EXTENSIONS):
                found.append(os.path.join(dirpath, fname))
    return sorted(found)


def is_blank_page(img, ink_fraction_threshold=BLANK_PAGE_INK_FRACTION_THRESHOLD):
    """Cheap, free (no API call) check for a blank/near-blank page."""
    try:
        gray = np.array(img.convert("L"))
        ink_pixels = np.count_nonzero(gray < 245)
        fraction = ink_pixels / gray.size
        return fraction < ink_fraction_threshold
    except Exception:
        return False


def load_pages(path, safe_print=print):
    ext = path.lower().rsplit(".", 1)[-1]
    if ext == "pdf":
        doc = fitz.open(path)
        pages = [Image.open(io.BytesIO(page.get_pixmap(dpi=200).tobytes("png"))) for page in doc]
        kept = []
        for i, img in enumerate(pages):
            if is_blank_page(img):
                safe_print(f"    [SKIP BLANK PAGE] {os.path.basename(path)} page {i + 1}")
                continue
            kept.append(img)
        return kept if kept else pages[:1]  # never return zero pages
    return [Image.open(path).convert("RGB")]


def ocr_page_data(img):
    """Return (text, mean_confidence) for a page using Tesseract's
    detailed output so we can decide detail level + hint inclusion."""
    try:
        data = pytesseract.image_to_data(img, config="--oem 3 --psm 6",
                                          output_type=pytesseract.Output.DICT)
        words, confs = [], []
        for word, conf in zip(data["text"], data["conf"]):
            word = word.strip()
            if not word:
                continue
            try:
                c = float(conf)
            except (TypeError, ValueError):
                continue
            if c < 0:
                continue
            words.append(word)
            confs.append(c)
        text = " ".join(words)
        mean_conf = sum(confs) / len(confs) if confs else 0.0
        return text, mean_conf
    except Exception as e:
        return f"[OCR unavailable: {e}]", 0.0


def ocr_hint_text(page_infos):
    """page_infos: list of dicts with 'text' and 'confidence'. Filters out
    low-value/noisy OCR (mostly from handwriting) and caps length to save
    prompt tokens."""
    chunks = []
    for i, info in enumerate(page_infos):
        text, conf = info["text"], info["confidence"]
        if conf < OCR_HINT_MIN_CONFIDENCE_TO_INCLUDE:
            chunks.append(
                f"--- OCR of page {i + 1}: low confidence ({conf:.0f}), "
                f"likely handwritten/noisy -- OCR hint omitted, rely on the image ---"
            )
            continue
        trimmed = text[:OCR_HINT_MAX_CHARS_PER_PAGE]
        if len(text) > OCR_HINT_MAX_CHARS_PER_PAGE:
            trimmed += " …[truncated]"
        chunks.append(
            f"--- OCR of page {i + 1} (typed text only, may contain errors,\n"
            f"    handwriting likely garbled -- use as a hint, not ground truth) ---\n{trimmed}"
        )
    return "\n\n".join(chunks)


def build_prompt(ocr_hint):
    return (
        f"{BASE_PROMPT}\n\n"
        f"For reference, here is raw OCR text pulled from this document. It's "
        f"machine-generated and may have errors -- especially anything "
        f"handwritten -- so treat it only as a hint for reading dense typed "
        f"sections, not as the final answer:\n\n{ocr_hint}\n\n"
        f"Now return ONLY the JSON array as instructed above."
    )


def resize_for_model(img, max_dim=MAX_IMAGE_DIMENSION):
    w, h = img.size
    longest = max(w, h)
    if longest <= max_dim:
        return img
    scale = max_dim / longest
    return img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)


def pil_to_data_uri(img):
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    b64 = base64.b64encode(buf.getvalue()).decode("utf-8")
    return f"data:image/png;base64,{b64}"


def choose_detail_level(ocr_text, ocr_confidence):
    """Adaptive cost control: cleanly-typed, dense pages -> 'low' detail
    (much cheaper). Sparse, low-confidence, or likely-handwritten pages
    keep 'high' detail so accuracy is preserved where it matters."""
    if ocr_confidence >= LOW_DETAIL_OCR_CONFIDENCE_THRESHOLD and \
       len(ocr_text) >= LOW_DETAIL_MIN_CHAR_COUNT:
        return "low"
    return "high"


# -----------------------------------------------------------------------
# JSON EXTRACTION / SCORING
# -----------------------------------------------------------------------

def extract_json(text):
    text = text.strip()
    text = re.sub(r"<think>.*?</think>", "", text, flags=re.DOTALL).strip()
    try:
        return json.loads(text)
    except Exception:
        pass
    fenced = re.search(r"```(?:json)?\s*(.*?)```", text, re.DOTALL)
    if fenced:
        try:
            return json.loads(fenced.group(1).strip())
        except Exception:
            pass
    start = text.find("[")
    if start != -1:
        depth = 0
        for i in range(start, len(text)):
            if text[i] == "[":
                depth += 1
            elif text[i] == "]":
                depth -= 1
                if depth == 0:
                    try:
                        return json.loads(text[start:i + 1])
                    except Exception:
                        break
    return None


def score_extraction(records):
    if not records:
        return -1
    score = 0
    for r in records:
        if not isinstance(r, dict):
            continue
        for col in ("date", "payer", "payee", "amount", "reference_no", "bank", "account_no"):
            if r.get(col, "") not in ("", None):
                score += 1
        amount = r.get("amount", "")
        if amount not in ("", None):
            try:
                float(str(amount).replace(",", "").strip())
                score += 2
            except ValueError:
                pass
        extra = r.get("extra_fields") or {}
        if isinstance(extra, dict):
            score += sum(1 for v in extra.values() if v not in ("", None))
    return score


# -----------------------------------------------------------------------
# API CALL
# -----------------------------------------------------------------------

def check_openai_key():
    if not OPENAI_API_KEY:
        print("ERROR: OPENAI_API_KEY environment variable is not set.\n"
              "  export OPENAI_API_KEY=\"sk-...\"   (then re-run)")
        return False
    return True


def _extract_single_sample(client, path, image_data_uris_with_detail, prompt,
                            debug_dir=None, safe_print=print, sample_label=""):
    tag = f" [{sample_label}]" if sample_label else ""
    content = [{"type": "text", "text": prompt}]
    for uri, detail in image_data_uris_with_detail:
        content.append({"type": "image_url", "image_url": {"url": uri, "detail": detail}})

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = client.chat.completions.create(
                model=OPENAI_MODEL,
                messages=[{"role": "user", "content": content}],
                temperature=MODEL_TEMPERATURE,
                max_tokens=MAX_OUTPUT_TOKENS,
                timeout=REQUEST_TIMEOUT,
            )
            raw = resp.choices[0].message.content or ""
            finish_reason = resp.choices[0].finish_reason
            parsed = extract_json(raw) if raw else None
            if parsed is not None:
                return parsed

            safe_print(f"    [{os.path.basename(path)}]{tag} "
                       f"[{'empty response' if not raw else 'parse failed'}, "
                       f"retry {attempt}/{MAX_RETRIES}] finish_reason={finish_reason}")
            if debug_dir and raw:
                os.makedirs(debug_dir, exist_ok=True)
                with open(os.path.join(debug_dir, f"raw_{sample_label}_attempt{attempt}.txt"),
                          "w", encoding="utf-8") as fh:
                    fh.write(raw)
            if finish_reason == "length":
                safe_print(f"    [{os.path.basename(path)}]{tag} "
                           f"[response truncated -- consider raising MAX_OUTPUT_TOKENS]")
        except Exception as e:
            safe_print(f"    [{os.path.basename(path)}]{tag} [retry {attempt}/{MAX_RETRIES}] {e}")
        time.sleep(5 * attempt)
    return None


def extract(client, path, images, debug_dir=None, safe_print=print):
    # OCR each page once; reuse for both the hint text and the detail-level decision
    page_infos = []
    for img in images:
        text, conf = ocr_page_data(img)
        page_infos.append({"text": text, "confidence": conf})

    ocr_hint = ocr_hint_text(page_infos)
    prompt = build_prompt(ocr_hint)

    image_data_uris_with_detail = []
    low_detail_count = 0
    for img, info in zip(images, page_infos):
        detail = choose_detail_level(info["text"], info["confidence"])
        if detail == "low":
            low_detail_count += 1
        resized = resize_for_model(img)
        image_data_uris_with_detail.append((pil_to_data_uri(resized), detail))

    if low_detail_count:
        safe_print(f"    [{os.path.basename(path)}] {low_detail_count}/{len(images)} "
                   f"page(s) sent at low detail (cleanly typed)")

    best_records, best_score = None, None
    for i in range(1, NUM_SAMPLES_PER_FILE + 1):
        label = f"sample {i}/{NUM_SAMPLES_PER_FILE}"
        records = _extract_single_sample(
            client, path, image_data_uris_with_detail, prompt,
            debug_dir=debug_dir, safe_print=safe_print, sample_label=label,
        )
        if records is None:
            continue
        s = score_extraction(records)
        safe_print(f"    [{os.path.basename(path)}] {label}: {len(records)} record(s), score={s}")
        if best_score is None or s > best_score:
            best_records, best_score = records, s

    if best_records is None:
        safe_print(f"    [FAILED] {os.path.basename(path)} -- skipped, needs manual entry")
        return []
    return best_records


# -----------------------------------------------------------------------
# FLATTENING / EXCEL WRITING  (unchanged from original)
# -----------------------------------------------------------------------

def flatten_records(results, source_file):
    core_rows = []
    for r in results:
        core = {col: r.get(col, "") for col in CORE_COLUMNS}
        core["source_file"] = source_file
        raw_extra = r.get("extra_fields") or {}
        norm_extra = {}
        if isinstance(raw_extra, dict):
            for k, v in raw_extra.items():
                nk = normalize_key(str(k))
                if nk:
                    norm_extra[nk] = json.dumps(v, ensure_ascii=False) if isinstance(v, (list, dict)) else v
        core["_extra"] = norm_extra
        core_rows.append(core)
    return core_rows


def build_extra_columns(core_rows):
    freq, order = {}, []
    for core in core_rows:
        for k in core.get("_extra", {}):
            if k not in freq:
                freq[k] = 0
                order.append(k)
            freq[k] += 1
    ranked = sorted(order, key=lambda k: (-freq[k], order.index(k)))
    kept = ranked[:MAX_EXTRA_COLUMNS]
    dropped = set(ranked[MAX_EXTRA_COLUMNS:])
    if dropped:
        for core in core_rows:
            extra = core.get("_extra", {})
            overflow = {k: v for k, v in extra.items() if k in dropped}
            if overflow:
                extra_str = "; ".join(f"{k}: {v}" for k, v in overflow.items())
                core["notes"] = (core["notes"] + "; " if core["notes"] else "") + extra_str
    return [k for k in order if k in set(kept)]


def style_workbook(path, amt_col_idx, n_core_cols):
    wb = load_workbook(path)
    ws = wb.active
    n_cols = ws.max_column

    header_fill = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
    extra_header_fill = PatternFill("solid", start_color="2E7D32", end_color="2E7D32")
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, n_cols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = extra_header_fill if col > n_core_cols else header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    for r in range(2, ws.max_row + 1):
        for col in range(1, n_cols + 1):
            c = ws.cell(row=r, column=col)
            c.font = Font(name="Arial", size=10)
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if col == amt_col_idx:
                c.number_format = "#,##0.00"
                c.alignment = Alignment(horizontal="right", vertical="center")

    core_widths = [30, 22, 14, 26, 26, 14, 36, 20, 18, 36]
    for i, w in enumerate(core_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for i in range(n_core_cols + 1, n_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18
    ws.freeze_panes = "A2"

    if ws.max_row > 1:
        last_row = ws.max_row
        total_row = last_row + 2
        col_letter = get_column_letter(amt_col_idx)
        label_col = max(1, amt_col_idx - 1)
        ws.cell(row=total_row, column=label_col, value="TOTAL").font = Font(name="Arial", bold=True)
        tc = ws.cell(row=total_row, column=amt_col_idx,
                     value=f"=SUM({col_letter}2:{col_letter}{last_row})")
        tc.number_format = "#,##0.00"
        tc.font = Font(name="Arial", bold=True)

    wb.save(path)


def write_excel(core_rows, extra_columns, out_path):
    all_columns = CORE_COLUMNS + extra_columns
    rows = []
    for core in core_rows:
        row = {col: core.get(col, "") for col in CORE_COLUMNS}
        extra = core.get("_extra", {})
        for ec in extra_columns:
            row[ec] = extra.get(ec, "")
        rows.append(row)

    df = pd.DataFrame(rows, columns=all_columns)

    def to_number(v):
        if v in ("", None):
            return None
        try:
            return float(str(v).replace(",", "").strip())
        except ValueError:
            return v

    df["amount"] = df["amount"].apply(to_number)
    for col in extra_columns:
        if col in NUMERIC_KEYS:
            df[col] = df[col].apply(to_number)

    try:
        df.to_excel(out_path, index=False)
        amt_col_idx = CORE_COLUMNS.index("amount") + 1
        style_workbook(out_path, amt_col_idx, len(CORE_COLUMNS))
    except PermissionError:
        raise PermissionError(
            f"'{out_path}' is open in Excel -- close it and retry."
        )


# -----------------------------------------------------------------------
# PER-FILE OUTPUT  (unchanged from original)
# -----------------------------------------------------------------------

def write_per_file_output(source_path, results, output_root, safe_print=print):
    base_name = os.path.splitext(os.path.basename(source_path))[0]
    ext       = os.path.splitext(source_path)[1]
    out_dir   = os.path.join(output_root, base_name)
    os.makedirs(out_dir, exist_ok=True)

    dest_image = os.path.join(out_dir, base_name + ext)
    try:
        shutil.copy2(source_path, dest_image)
    except Exception as e:
        safe_print(f"    [WARNING] could not copy source file: {e}")

    if not results:
        safe_print(f"    [WARNING] no records for '{base_name}' -- Excel not written.")
        return

    core_rows     = flatten_records(results, os.path.basename(source_path))
    extra_columns = build_extra_columns(core_rows)
    xlsx_path     = os.path.join(out_dir, base_name + ".xlsx")
    try:
        write_excel(core_rows, extra_columns, xlsx_path)
        safe_print(f"    -> {xlsx_path}  ({len(core_rows)} record(s))")
    except PermissionError as e:
        safe_print(f"    [SKIPPED WRITE] {e}")


# -----------------------------------------------------------------------
# STATE MANAGEMENT  (unchanged from original)
# -----------------------------------------------------------------------

def state_file_path(output_root):
    return os.path.join(output_root, STATE_FILE_NAME)


def load_state(output_root):
    path = state_file_path(output_root)
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as fh:
                return json.load(fh)
        except json.JSONDecodeError:
            print(f"    [WARNING] {path} is corrupted -- starting fresh state.")
    return {"processed_files": []}


def save_state(state, output_root):
    path = state_file_path(output_root)
    tmp  = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as fh:
        json.dump(state, fh, ensure_ascii=False, indent=2)
    os.replace(tmp, path)


# -----------------------------------------------------------------------
# MAIN CONTINUOUS LOOP  (unchanged from original)
# -----------------------------------------------------------------------

_shutdown_requested = threading.Event()


def _handle_sigint(signum, frame):
    print("\n\nCtrl+C received -- finishing in-flight files then stopping. "
          "Press Ctrl+C again to force-quit immediately.")
    _shutdown_requested.set()


def discover_input_folders(base_dir):
    result = []
    try:
        entries = os.listdir(base_dir)
    except OSError:
        return result
    for name in sorted(entries):
        if name.startswith(".") or name.startswith("_"):
            continue
        if name.endswith(OUTPUT_SUFFIX):
            continue
        full = os.path.join(base_dir, name)
        if os.path.isdir(full):
            result.append(full)
    return result


def output_root_for(input_folder):
    name   = os.path.basename(input_folder.rstrip(os.sep))
    parent = os.path.dirname(input_folder.rstrip(os.sep))
    return os.path.join(parent, name + OUTPUT_SUFFIX)


def main():
    print("=" * 60)
    print("CONFIG (cost-optimized):")
    print(f"  OPENAI_MODEL           = {OPENAI_MODEL}")
    print(f"  API_CONCURRENT_WORKERS = {API_CONCURRENT_WORKERS}  (per input folder)")
    print(f"  NUM_SAMPLES_PER_FILE   = {NUM_SAMPLES_PER_FILE}")
    print(f"  MAX_RETRIES            = {MAX_RETRIES}")
    print(f"  POLL_INTERVAL_SECONDS  = {POLL_INTERVAL_SECONDS}")
    print(f"  Adaptive image detail  = low-detail for typed pages "
          f"(conf>={LOW_DETAIL_OCR_CONFIDENCE_THRESHOLD}, chars>={LOW_DETAIL_MIN_CHAR_COUNT})")
    print(f"  OCR hint cap           = {OCR_HINT_MAX_CHARS_PER_PAGE} chars/page, "
          f"omitted below conf {OCR_HINT_MIN_CONFIDENCE_TO_INCLUDE}")
    print(f"  Blank page skipping    = ink fraction < {BLANK_PAGE_INK_FRACTION_THRESHOLD}")
    print(f"  BASE_DIR               = {BASE_DIR}  (scanned for input folders)")
    print(f"  OUTPUT_SUFFIX          = {OUTPUT_SUFFIX}")
    print("  OUTPUT: <InputFolder>{OUTPUT_SUFFIX}/<basename>/  ->  <basename>.xlsx + copy of source")
    print("  MODE: continuous -- runs until Ctrl+C")
    print("=" * 60 + "\n")

    if not check_openai_key():
        return

    client = OpenAI(api_key=OPENAI_API_KEY)

    signal.signal(signal.SIGINT, _handle_sigint)

    failed       = []
    failed_lock  = threading.Lock()
    print_lock   = threading.Lock()

    def safe_print(msg):
        with print_lock:
            print(msg)

    folder_states     = {}
    folder_locks      = {}
    folders_meta_lock = threading.Lock()

    def get_folder_context(output_root):
        with folders_meta_lock:
            if output_root not in folder_states:
                os.makedirs(output_root, exist_ok=True)
                st = load_state(output_root)
                folder_states[output_root] = st
                folder_locks[output_root]  = threading.Lock()
                if st["processed_files"]:
                    safe_print(
                        f"[{os.path.basename(output_root)}] Resuming: "
                        f"{len(st['processed_files'])} already done."
                    )
            return folder_states[output_root], folder_locks[output_root]

    def process_one(f, output_root):
        base_name  = os.path.splitext(os.path.basename(f))[0]
        folder_tag = os.path.basename(output_root)

        out_dir = os.path.join(output_root, base_name)
        os.makedirs(out_dir, exist_ok=True)

        try:
            images  = load_pages(f, safe_print=safe_print)
            results = extract(client, f, images, debug_dir=out_dir, safe_print=safe_print)

            cp  = os.path.join(out_dir, "raw_extraction.json")
            tmp = cp + ".tmp"
            with open(tmp, "w", encoding="utf-8") as fh:
                json.dump(results, fh, ensure_ascii=False, indent=2)
            os.replace(tmp, cp)

            write_per_file_output(f, results, output_root, safe_print=safe_print)

            safe_print(f"[OK][{folder_tag}] {os.path.basename(f)} -> {len(results)} record(s)")

            state, lock = get_folder_context(output_root)
            with lock:
                if base_name not in state["processed_files"]:
                    state["processed_files"].append(base_name)
                save_state(state, output_root)

        except Exception as e:
            safe_print(f"[SKIPPED - error][{folder_tag}] {os.path.basename(f)}: {e}")
            with failed_lock:
                failed.append(f"[{folder_tag}] {os.path.basename(f)}")

    safe_print("Watching for input folders and files... (Ctrl+C to stop)\n")

    while not _shutdown_requested.is_set():
        input_folders = discover_input_folders(BASE_DIR)

        if not input_folders:
            safe_print("  No input folders found yet -- waiting...")
        else:
            all_new = []
            for input_folder in input_folders:
                output_root     = output_root_for(input_folder)
                state, _        = get_folder_context(output_root)
                already_done    = set(state["processed_files"])
                for f in discover_input_files(input_folder):
                    if os.path.splitext(os.path.basename(f))[0] not in already_done:
                        all_new.append((f, output_root))

            if all_new:
                safe_print(
                    f"Found {len(all_new)} new file(s) across "
                    f"{len(input_folders)} folder(s): "
                    + ", ".join(os.path.basename(d) for d in input_folders)
                )
                with ThreadPoolExecutor(
                    max_workers=API_CONCURRENT_WORKERS * len(input_folders)
                ) as executor:
                    futures = [executor.submit(process_one, f, out) for f, out in all_new]
                    for _ in as_completed(futures):
                        if _shutdown_requested.is_set():
                            break

        if _shutdown_requested.is_set():
            break

        time.sleep(POLL_INTERVAL_SECONDS)

    if failed:
        safe_print(
            f"\n{len(failed)} file(s) failed and will be retried next run:\n"
            + "\n".join(f"  {x}" for x in failed)
        )

    safe_print("\nStopped.")


if __name__ == "__main__":
    main()
