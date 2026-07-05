"""
Invoice Extraction Pipeline -- Gemini (free tier), with dynamic fields
------------------------------------------------------------------------------
Same job as the plain version, but no longer forces every document into one
rigid column layout. Real invoices/cheques/vouchers vary a lot (discount,
GST, freight, PO number, due date, TDS, etc.), so:

  - CORE fields (document_type, date, payer, payee, amount, reference_no,
    bank, account_no, notes) are always extracted and always appear as the
    first columns -- these exist on almost every document type here.
  - Anything else the model finds on a document (discount, tax, GST number,
    due date, PO number, freight, TDS, cheque validity date, etc.) goes into
    an "extra_fields" object. Per OUTPUT FILE, all extra keys seen across
    that file's records are unioned into extra columns appended after the
    core ones -- so a file with 3 cheques and 1 GST invoice gets one sheet
    with "Discount"/"GST No" columns filled only where relevant, blank
    elsewhere, instead of losing that data or forcing a one-size-fits-all
    schema across your whole batch.

Also, like the Ollama version:
  - RESULTS ARE CACHED per file (raw_extraction.json) -- re-running the
    script never re-calls the (rate-limited, quota-limited) Gemini API for
    a file that's already been extracted. Only the Excel layout gets
    rebuilt every run, so newly-seen extra fields from later files still
    show up as columns in earlier files' sheets too.
  - A combined "All_Invoices_Combined.xlsx" master sheet with every record
    from every file.
  - A "All_Invoices_Visual_Report.xlsx" with a thumbnail of each source
    image next to a table of just that invoice's own fields.

MODEL FALLBACK:
  gemini-2.5-flash-lite is tried first (best free-tier quota). If it fails
  all its retries on a given file (parse errors, transient 503s, etc.) the
  script automatically retries that same file with gemini-2.5-flash before
  giving up on it. If flash-lite's daily quota runs out entirely, the rest
  of the run switches to flash-2.5 automatically -- the whole batch only
  stops if BOTH models are out of quota.

SETUP:
    pip install google-genai pymupdf pillow pandas openpyxl
    export GEMINI_API_KEY="your-key-here"      # https://aistudio.google.com/apikey

USAGE:
    python invoice_pipeline_simple.py

    (reads from ./input_img, writes to ./invoice_output/<filename>/)
"""

import os
import io
import re
import json
import time
import glob
import shutil
import tempfile

from google import genai
from google.genai import types
from PIL import Image
import fitz  # PyMuPDF
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- CONFIG ----------
INPUT_DIR = "IV"
OUTPUT_ROOT = "invoice_output2"

# gemini-2.5-flash-lite has the free tier's most generous quota (~15 RPM /
# ~1000 RPD as of mid-2026) and is plenty accurate for structured extraction.
# gemini-2.5-flash is the fallback: tried per-file if flash-lite fails all
# its retries (parse errors, transient errors), and used for the rest of
# the run if flash-lite's daily quota runs out entirely -- so the whole
# batch doesn't grind to a halt just because one model's free quota is spent.
GEMINI_MODEL = "gemini-2.5-flash-lite"
GEMINI_MODEL_FALLBACK = "gemini-2.5-flash"
GEMINI_MODELS = [GEMINI_MODEL, GEMINI_MODEL_FALLBACK]

# 5s between calls = 12 requests/minute, safely under the ~15 RPM free cap
# (leaves headroom for the occasional retry without tripping the limit).
SLEEP_BETWEEN_CALLS = 5

# Self-imposed daily cap, kept comfortably below the real ~1000 RPD free
# quota. Check your live number at https://aistudio.google.com/rate-limit
# and adjust if needed.
DAILY_LIMIT = 900

MAX_RETRIES = 4
FILE_LOCK_RETRIES = 5     # extra retries specifically for a locked output .xlsx
FILE_LOCK_WAIT = 15       # seconds between locked-file retries

# Cap on how many distinct extra-field columns a single file's sheet can
# grow to. Guards against a garbled model response spraying dozens of junk
# keys into the header row. The exhaustive prompt below routinely surfaces
# 8-15 genuinely distinct fields (time, currency, subtotal, tax,
# payment_method, rep, addresses, etc.) even on a single simple receipt.
MAX_EXTRA_COLUMNS = 25

# Keys (after normalize_key) that should be written as real numbers rather
# than text, so Excel SUM()/number formatting works on them instead of
# silently treating them as blank.
NUMERIC_KEYS = {"subtotal", "tax", "discount", "tds", "freight",
                 "balance_remaining"}

# Visual report layout: thumbnail width in pixels, blank rows between one
# invoice's block (image + table) and the next, and the row-height estimate
# used to reserve enough rows under an image before the table starts.
VISUAL_THUMB_MAX_DIM = 220
VISUAL_GAP_ROWS = 4
EXCEL_DEFAULT_ROW_HEIGHT_PX = 20  # ~15pt default row height in pixels
# -----------------------------

CORE_COLUMNS = ["source_file", "document_type", "date", "payer", "payee",
                "amount", "reference_no", "bank", "account_no", "notes"]

# Known synonyms get normalized to one canonical key so "Discount %" and
# "discount_percent" don't become two separate columns. Match is
# case/whitespace-insensitive against the lowercased, underscored key the
# model returns.
FIELD_ALIASES = {
    "discount": "discount",
    "discount_amount": "discount",
    "discount_percent": "discount",
    "gst": "gst_no",
    "gst_no": "gst_no",
    "gst_number": "gst_no",
    "gstin": "gst_no",
    "tax": "tax",
    "tax_amount": "tax",
    "tds": "tds",
    "due_date": "due_date",
    "payment_due_date": "due_date",
    "po_number": "po_number",
    "purchase_order_no": "po_number",
    "freight": "freight",
    "shipping_charges": "freight",
    "cheque_validity_date": "cheque_validity_date",
    "ifsc_code": "ifsc_code",
    "ifsc": "ifsc_code",
}


def normalize_key(k):
    k = re.sub(r"[^a-z0-9]+", "_", k.strip().lower()).strip("_")
    return FIELD_ALIASES.get(k, k)


class QuotaExhausted(Exception):
    """Raised when Gemini reports the daily (RPD) quota is used up -- no point retrying."""
    pass


BASE_PROMPT = """Extract structured data from this Indian bank/advertising payment
document. It may contain a MIX of typed/printed text and handwritten text in
the same image (e.g. a printed form filled in by hand, or a receipt with a
handwritten section next to a printed letterhead). Read both carefully,
using the actual image as ground truth.

Return a JSON array. Each element = ONE distinct financial transaction or
record found anywhere in the image (a single image may show multiple cheques,
receipts, or line-items -- list each one separately, do not merge them).

For each record extract these CORE fields (always include all of them, use ""
if missing/illegible):
- document_type: e.g. "Cheque", "Deposit Slip", "NEFT Advice", "Payment Voucher",
  "Release Order", "Receipt", "Tax Invoice", "Invoice"
- date: exactly as written (date only -- put any separate time value in extra_fields, see below)
- payer: who is paying / sending money / the client being billed
- payee: who receives the money / the business issuing the document
- amount: the FINAL/net payable or total amount, number only, no currency
  symbols, no commas, no units (e.g. "540" not "USD540" or "540.00 USD")
- reference_no: cheque no. / UTR / transaction ID / receipt no. / invoice no. / bill no. --
  whichever is present (combine if multiple, comma separated)
- bank: bank name involved (leave "" if not a banking document)
- account_no: as shown, including masked digits if partially hidden
- notes: only for genuinely unstructured remarks that don't fit as a labeled
  field anywhere else (e.g. a free-text comment) -- do NOT use notes as a
  dumping ground for labeled fields, those belong in extra_fields below

CRITICAL -- BE EXHAUSTIVE: this document may be one of thousands of
different formats (invoices, receipts, cheques, vouchers, advices) and each
one can carry different labeled information. Do not limit yourself to a
fixed list. Scan the ENTIRE document top to bottom and capture EVERY
distinct labeled piece of information you can see as a key in
"extra_fields", even if it seems minor or you're unsure it matters.
Concretely, always check for (include only if actually present on the
document):
- time (if a timestamp with a time component appears, separate from date)
- currency / currency_unit (e.g. "USD", "INR") if the amount has one
- subtotal, tax, discount, gst_no, tds, freight, balance_remaining,
  amount_in_words
- invoice_no (if distinct from reference_no), po_number, due_date,
  payment_method, rep / prepared_by / issued_by, client_address,
  company_address, company_phone, company_email, ifsc_code,
  branch, cheque_validity_date
- ANY other labeled field visible on the document that isn't covered above
  -- invent a short snake_case key for it rather than skipping it or
  folding it into notes.

Use short snake_case keys and plain values -- strip currency symbols/commas
from numeric-looking values (put the currency itself in a separate
"currency" key if shown). Only include keys for data actually present in
THIS document -- do not invent data, and leave "extra_fields" as {} only if
there is truly nothing beyond the core fields.

Return ONLY the JSON array -- no markdown fences, no explanation, no
commentary before or after the array. Example shape:
[{"document_type": "...", "date": "...", "payer": "...", "payee": "...",
  "amount": "540", "reference_no": "...", "bank": "...", "account_no": "...",
  "notes": "...", "extra_fields": {"time": "11:10", "currency": "USD",
  "subtotal": "540", "tax": "0", "balance_remaining": "0",
  "payment_method": "Bank Transfer", "rep": "J. Nolan"}}]"""


def load_pages(path):
    ext = path.lower().rsplit(".", 1)[-1]
    if ext == "pdf":
        doc = fitz.open(path)
        return [Image.open(io.BytesIO(page.get_pixmap(dpi=200).tobytes("png"))) for page in doc]
    return [Image.open(path).convert("RGB")]


def extract_json(text):
    text = text.strip()

    # Some Gemini responses (or future model swaps) may wrap the answer in
    # a reasoning block or code fence -- strip/locate the JSON robustly
    # instead of assuming clean output.
    text = re.sub(r"<think>.*?</think>", "", text, flags=re.DOTALL).strip()

    try:
        return json.loads(text)
    except Exception:
        pass

    fenced = re.search(r"```(?:json)?\s*(.*?)```", text, re.DOTALL)
    if fenced:
        try:
            return json.loads(fenced.group(1).strip())
        except Exception:
            pass

    start = text.find("[")
    if start != -1:
        depth = 0
        for i in range(start, len(text)):
            if text[i] == "[":
                depth += 1
            elif text[i] == "]":
                depth -= 1
                if depth == 0:
                    candidate = text[start:i + 1]
                    try:
                        return json.loads(candidate)
                    except Exception:
                        break

    return None


# Models whose daily quota has been confirmed exhausted THIS RUN. Tracked
# globally (not per-file) so once flash-lite's RPD limit is hit on file #5,
# files #6+ don't waste a full MAX_RETRIES cycle re-discovering that --
# they skip straight to the fallback model.
_exhausted_models = set()


def extract(path, client, debug_dir=None):
    images = load_pages(path)
    content = [BASE_PROMPT] + images

    for model_name in GEMINI_MODELS:
        if model_name in _exhausted_models:
            continue

        for attempt in range(1, MAX_RETRIES + 1):
            try:
                resp = client.models.generate_content(
                    model=model_name,
                    contents=content,
                    config=types.GenerateContentConfig(
                        response_mime_type="application/json",
                        temperature=0.1,
                    ),
                )
                raw = resp.text or ""
                parsed = extract_json(raw)
                if parsed is not None:
                    return parsed
                print(f"    [parse failed, retry {attempt}/{MAX_RETRIES}] "
                      f"{model_name} didn't return valid JSON")
                if debug_dir:
                    os.makedirs(debug_dir, exist_ok=True)
                    dump_path = os.path.join(
                        debug_dir, f"raw_response_{model_name}_attempt{attempt}.txt")
                    with open(dump_path, "w", encoding="utf-8") as fh:
                        fh.write(raw)
                    print(f"        raw response saved to {dump_path}")

            except Exception as e:
                msg = str(e)

                # RPD (daily quota) exhaustion won't fix itself with a
                # retry -- mark this model exhausted and drop straight
                # through to the next model in GEMINI_MODELS (if any),
                # rather than burning the remaining retries on it.
                if "RESOURCE_EXHAUSTED" in msg or "429" in msg:
                    if "PerDay" in msg or "per day" in msg.lower() or "daily" in msg.lower():
                        print(f"    [{model_name} daily quota exhausted]")
                        _exhausted_models.add(model_name)
                        break
                    print(f"    [rate limited on {model_name}, backing off] "
                          f"attempt {attempt}/{MAX_RETRIES}")
                    time.sleep(20 * attempt)
                    continue

                # 503 UNAVAILABLE (Google's servers busy) -- transient,
                # worth a longer backoff before retrying.
                if "UNAVAILABLE" in msg or "503" in msg:
                    print(f"    [{model_name} busy, waiting] attempt {attempt}/{MAX_RETRIES}")
                    time.sleep(15 * attempt)
                    continue

                print(f"    [retry {attempt}/{MAX_RETRIES} on {model_name}] {e}")
                time.sleep(5 * attempt)

        # This model's retries are used up (or its quota just hit) with no
        # successful parse -- fall through to the next model in the list,
        # printing which one we're switching to, if there is one left.
        remaining = [m for m in GEMINI_MODELS if m != model_name and m not in _exhausted_models]
        if remaining:
            print(f"    [falling back from {model_name} to {remaining[0]}]")

    # Only stop the whole run if EVERY configured model is quota-exhausted --
    # a plain parse/transient failure on all models just fails this one file.
    if set(GEMINI_MODELS) <= _exhausted_models:
        raise QuotaExhausted("All configured Gemini models have hit their daily quota.")

    print(f"    [FAILED] {os.path.basename(path)} -- skipped, needs manual entry")
    return []


def flatten_records(results, source_file):
    """Split each raw model record into its core dict + normalized
    extra_fields dict. No capping here -- this file's fields get unioned
    with every other file's fields first, and the cap (if any) is applied
    once globally in main() so column layout stays consistent batch-wide."""
    core_rows = []

    for r in results:
        core = {col: r.get(col, "") for col in CORE_COLUMNS}
        core["source_file"] = source_file

        raw_extra = r.get("extra_fields") or {}
        norm_extra = {}
        if isinstance(raw_extra, dict):
            for k, v in raw_extra.items():
                nk = normalize_key(str(k))
                if nk:
                    norm_extra[nk] = v

        core["_extra"] = norm_extra
        core_rows.append(core)

    return core_rows


def build_global_extra_columns(all_core_rows):
    """Union extra_fields keys across EVERY row in the whole batch, in
    first-seen order. Cap applied once, globally -- fields beyond the cap
    fold into that row's notes instead of getting a column, so a rare
    one-off field on invoice #847 doesn't blow up every sheet's width, but
    common fields shared across many invoices still get their own column
    no matter how far into the batch they first appear."""
    freq = {}
    order = []
    for core in all_core_rows:
        for k in core.get("_extra", {}):
            if k not in freq:
                freq[k] = 0
                order.append(k)
            freq[k] += 1

    ranked = sorted(order, key=lambda k: (-freq[k], order.index(k)))
    kept = ranked[:MAX_EXTRA_COLUMNS]
    dropped = set(ranked[MAX_EXTRA_COLUMNS:])

    if dropped:
        for core in all_core_rows:
            extra = core.get("_extra", {})
            overflow = {k: v for k, v in extra.items() if k in dropped}
            if overflow:
                extra_str = "; ".join(f"{k}: {v}" for k, v in overflow.items())
                core["notes"] = (core["notes"] + "; " if core["notes"] else "") + extra_str

    return [k for k in order if k in set(kept)]


def style_workbook(path, amt_col_letter, n_core_cols, n_extra_cols):
    wb = load_workbook(path)
    ws = wb.active
    n_cols = ws.max_column

    header_fill = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
    extra_header_fill = PatternFill("solid", start_color="2E7D32", end_color="2E7D32")
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, n_cols + 1):
        c = ws.cell(row=1, column=col)
        # Extra (dynamic) columns get a different header color so it's
        # visually obvious which fields are core vs. document-specific.
        c.fill = extra_header_fill if col > n_core_cols else header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    for r in range(2, ws.max_row + 1):
        for col in range(1, n_cols + 1):
            c = ws.cell(row=r, column=col)
            c.font = Font(name="Arial", size=10)
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if col == amt_col_letter:
                c.number_format = "#,##0.00"
                c.alignment = Alignment(horizontal="right", vertical="center")

    core_widths = [30, 22, 14, 26, 26, 14, 36, 20, 18, 36]
    for i, w in enumerate(core_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for i in range(n_core_cols + 1, n_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18
    ws.freeze_panes = "A2"

    if ws.max_row > 1:
        last_row = ws.max_row
        total_row = last_row + 2
        col_letter = get_column_letter(amt_col_letter)
        label_col = max(1, amt_col_letter - 1)
        ws.cell(row=total_row, column=label_col, value="TOTAL").font = Font(name="Arial", bold=True)
        tc = ws.cell(row=total_row, column=amt_col_letter,
                     value=f"=SUM({col_letter}2:{col_letter}{last_row})")
        tc.number_format = "#,##0.00"
        tc.font = Font(name="Arial", bold=True)

    wb.save(path)


def write_visual_report(files, per_file_rows, out_path):
    """Combined workbook where each invoice gets its own block: a thumbnail
    of the source image, then a table built from ONLY that invoice's own
    fields (core fields that had a value + whatever extra_fields it had) --
    not the global batch-wide column set."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    header_fill = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    filename_font = Font(name="Arial", bold=True, size=12)
    cell_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row_cursor = 1
    max_cols_used = 1
    tmp_dir = tempfile.mkdtemp(prefix="invoice_thumbs_")

    try:
        for f in files:
            base_name = os.path.splitext(os.path.basename(f))[0]
            rows = per_file_rows.get(base_name)
            if not rows:
                continue  # failed extraction, nothing to show for this file

            block_start_row = row_cursor

            # --- thumbnail ---
            img_rows_used = 6  # fallback if thumbnail fails to load
            try:
                pages = load_pages(f)
                thumb = pages[0].copy()
                thumb.thumbnail((VISUAL_THUMB_MAX_DIM, VISUAL_THUMB_MAX_DIM), Image.LANCZOS)
                thumb_path = os.path.join(tmp_dir, f"{base_name}.png")
                thumb.save(thumb_path)
                xl_img = XLImage(thumb_path)
                ws.add_image(xl_img, f"A{row_cursor}")
                img_rows_used = max(6, (xl_img.height // EXCEL_DEFAULT_ROW_HEIGHT_PX) + 2)
            except Exception as e:
                ws.cell(row=row_cursor, column=1, value=f"[thumbnail unavailable: {e}]").font = cell_font

            table_row = block_start_row + img_rows_used

            # --- filename label ---
            ws.cell(row=table_row, column=1, value=os.path.basename(f)).font = filename_font
            table_row += 1

            # --- this invoice's own columns, built from only what's present ---
            for rec in rows:
                cols = [c for c in CORE_COLUMNS if c != "source_file" and rec.get(c, "")]
                extra = rec.get("_extra", {})
                extra_cols = [k for k, v in extra.items() if v not in ("", None)]
                cols += extra_cols
                if not cols:
                    continue
                max_cols_used = max(max_cols_used, len(cols))

                for j, c in enumerate(cols, start=1):
                    hc = ws.cell(row=table_row, column=j, value=c.replace("_", " ").title())
                    hc.font = header_font
                    hc.fill = header_fill
                    hc.border = border
                    hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                table_row += 1

                for j, c in enumerate(cols, start=1):
                    val = rec.get(c, "") if c in CORE_COLUMNS else extra.get(c, "")
                    vc = ws.cell(row=table_row, column=j, value=val)
                    vc.font = cell_font
                    vc.border = border
                    vc.alignment = Alignment(vertical="center", wrap_text=True)
                table_row += 1

                table_row += 1  # small gap between records within the same file

            row_cursor = max(table_row, block_start_row + img_rows_used + 2) + VISUAL_GAP_ROWS

        for i in range(1, max_cols_used + 1):
            ws.column_dimensions[get_column_letter(i)].width = 20

        wb.save(out_path)
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


def write_master_excel(all_core_rows, extra_columns, out_path):
    """One combined sheet with every record from every file in the batch."""
    write_excel(all_core_rows, extra_columns, out_path)


def write_excel(core_rows, extra_columns, out_path):
    """Build one sheet: core columns first, then whatever extra fields
    actually showed up across this file's records (union, first-seen
    order). Cells left blank where a given record didn't have that field.
    Retries on PermissionError (file locked, almost always because it's
    open in Excel) instead of crashing the run."""
    all_columns = CORE_COLUMNS + extra_columns
    rows = []
    for core in core_rows:
        row = {col: core.get(col, "") for col in CORE_COLUMNS}
        extra = core.get("_extra", {})
        for ec in extra_columns:
            row[ec] = extra.get(ec, "")
        rows.append(row)

    df = pd.DataFrame(rows, columns=all_columns)

    def to_number(v):
        if v in ("", None):
            return None
        try:
            return float(str(v).replace(",", "").strip())
        except ValueError:
            return v

    df["amount"] = df["amount"].apply(to_number)
    for col in extra_columns:
        if col in NUMERIC_KEYS:
            df[col] = df[col].apply(to_number)

    amt_col_idx = CORE_COLUMNS.index("amount") + 1

    for attempt in range(1, FILE_LOCK_RETRIES + 1):
        try:
            df.to_excel(out_path, index=False)
            style_workbook(out_path, amt_col_idx, len(CORE_COLUMNS), len(extra_columns))
            return
        except PermissionError:
            print(f"    [locked] '{out_path}' is open in Excel (or similar) -- "
                  f"close it. Retrying in {FILE_LOCK_WAIT}s "
                  f"({attempt}/{FILE_LOCK_RETRIES})...")
            time.sleep(FILE_LOCK_WAIT)

    raise PermissionError(f"Could not write '{out_path}' after {FILE_LOCK_RETRIES} attempts -- "
                           f"file appears to still be open elsewhere.")


def main():
    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        print("ERROR: GEMINI_API_KEY is not set in this terminal session.")
        print("  Windows PowerShell (current session only): $env:GEMINI_API_KEY = \"your-key-here\"")
        print("  Windows (persist across sessions, then OPEN A NEW terminal): setx GEMINI_API_KEY \"your-key-here\"")
        print("  Mac/Linux: export GEMINI_API_KEY=\"your-key-here\"")
        return
    client = genai.Client(api_key=api_key)

    print("=" * 60)
    print("CONFIG IN USE (check this matches what you expect):")
    print(f"  GEMINI_MODEL        = {GEMINI_MODEL} (fallback: {GEMINI_MODEL_FALLBACK})")
    print(f"  DAILY_LIMIT         = {DAILY_LIMIT}")
    print(f"  MAX_EXTRA_COLUMNS   = {MAX_EXTRA_COLUMNS}")
    print("=" * 60 + "\n")

    files = sorted(
        f for f in glob.glob(os.path.join(INPUT_DIR, "*"))
        if f.lower().endswith((".jpg", ".jpeg", ".png", ".pdf"))
    )
    if not files:
        print(f"No .jpg/.jpeg/.png/.pdf files found in '{INPUT_DIR}/'")
        return

    os.makedirs(OUTPUT_ROOT, exist_ok=True)

    def raw_cache_path(f):
        base_name = os.path.splitext(os.path.basename(f))[0]
        return os.path.join(OUTPUT_ROOT, base_name, "raw_extraction.json")

    def already_extracted(f):
        # Cache the MODEL OUTPUT, not the final xlsx -- the xlsx layout gets
        # rebuilt from cache every run to pick up new columns from newer
        # files, but re-calling the (rate-limited, quota-limited) API on
        # unchanged files is wasted work (and burns your daily quota).
        return os.path.exists(raw_cache_path(f))

    pending = [f for f in files if not already_extracted(f)]
    cached_count = len(files) - len(pending)
    if cached_count:
        print(f"{cached_count} file(s) already extracted (using cached results).")
    if not pending:
        print(f"Nothing left to extract -- all {len(files)} file(s) already cached.")
    else:
        print(f"{len(pending)} file(s) to extract with {GEMINI_MODEL} "
              f"(falls back to {GEMINI_MODEL_FALLBACK} if needed). "
              f"Processing up to {DAILY_LIMIT} of them today.\n")

    failed = []
    processed_this_run = 0
    quota_hit = False

    # ---- PASS 1: extract (or load from cache) every file's raw records ----
    for f in pending:
        if processed_this_run >= DAILY_LIMIT:
            remaining = len(pending) - processed_this_run
            days_left = -(-remaining // DAILY_LIMIT)  # ceil division
            print(f"\nReached today's limit of {DAILY_LIMIT}. "
                  f"{remaining} file(s) left -- re-run tomorrow "
                  f"(~{days_left} more day(s) at this rate).")
            break

        base_name = os.path.splitext(os.path.basename(f))[0]
        print(f"[{processed_this_run + 1}/{min(len(pending), DAILY_LIMIT)}] Extracting {os.path.basename(f)}")

        folder = os.path.join(OUTPUT_ROOT, base_name)
        os.makedirs(folder, exist_ok=True)

        try:
            shutil.copy2(f, os.path.join(folder, os.path.basename(f)))
            results = extract(f, client, debug_dir=folder)
            tmp_path = raw_cache_path(f) + ".tmp"
            with open(tmp_path, "w", encoding="utf-8") as fh:
                json.dump(results, fh, ensure_ascii=False, indent=2)
            os.replace(tmp_path, raw_cache_path(f))  # atomic -- avoids a
            # half-written cache file if the script is killed mid-write
            print(f"    -> {len(results)} record(s) extracted")
            processed_this_run += 1

        except QuotaExhausted:
            remaining = len(pending) - processed_this_run
            print(f"\nDaily quota exhausted after {processed_this_run} file(s) this run. "
                  f"{remaining} file(s) left -- re-run once quota resets "
                  f"(resets at midnight Pacific time).")
            quota_hit = True
            break

        except Exception as e:
            print(f"    [SKIPPED - error] {os.path.basename(f)}: {e}")
            failed.append(os.path.basename(f))

        time.sleep(SLEEP_BETWEEN_CALLS)

    # ---- PASS 2: rebuild every sheet from cache with a consistent, ----
    # ---- batch-wide column layout (covers cached files from earlier runs too) ----
    print("\nBuilding spreadsheets with a consistent column layout across the batch...")

    per_file_rows = {}  # base_name -> core_rows (with _extra intact)
    for f in files:
        cache_path = raw_cache_path(f)
        if not os.path.exists(cache_path):
            continue  # this file failed extraction (or wasn't reached yet) -- no cache
        base_name = os.path.splitext(os.path.basename(f))[0]
        try:
            with open(cache_path, "r", encoding="utf-8") as fh:
                results = json.load(fh)
        except json.JSONDecodeError:
            print(f"    [BAD CACHE] {base_name}: raw_extraction.json is corrupted, "
                  f"deleting it -- re-run the script to re-extract this file.")
            os.remove(cache_path)
            failed.append(os.path.basename(f))
            continue
        per_file_rows[base_name] = flatten_records(results, os.path.basename(f))

    all_core_rows = [row for rows in per_file_rows.values() for row in rows]
    if not all_core_rows:
        print("No successfully extracted records to write.")
    else:
        extra_columns = build_global_extra_columns(all_core_rows)
        print(f"Global extra columns ({len(extra_columns)}): {', '.join(extra_columns) if extra_columns else '(none)'}")

        for base_name, rows in per_file_rows.items():
            out_xlsx = os.path.join(OUTPUT_ROOT, base_name, "Extracted_Invoice.xlsx")
            try:
                write_excel(rows, extra_columns, out_xlsx)
            except PermissionError as e:
                print(f"    [SKIPPED WRITE] {base_name}: {e}")

        master_path = os.path.join(OUTPUT_ROOT, "All_Invoices_Combined.xlsx")
        try:
            write_master_excel(all_core_rows, extra_columns, master_path)
            print(f"\nMaster combined workbook: {master_path} ({len(all_core_rows)} record(s) across {len(per_file_rows)} file(s))")
        except PermissionError as e:
            print(f"    [SKIPPED WRITE] master workbook: {e}")

        visual_path = os.path.join(OUTPUT_ROOT, "All_Invoices_Visual_Report.xlsx")
        try:
            write_visual_report(files, per_file_rows, visual_path)
            print(f"Visual report (thumbnail + per-invoice table): {visual_path}")
        except PermissionError:
            print(f"    [SKIPPED WRITE] visual report: '{visual_path}' is open elsewhere -- close it and re-run.")
        except Exception as e:
            print(f"    [SKIPPED WRITE] visual report failed: {e}")

    print(f"\nDone. Per-file results in '{OUTPUT_ROOT}/<filename>/Extracted_Invoice.xlsx'.")
    if failed:
        print(f"{len(failed)} file(s) failed extraction (re-run the script "
              f"to retry them automatically): {', '.join(failed)}")
    if quota_hit:
        print("Re-run the script after your Gemini quota resets to pick up where you left off.")


if __name__ == "__main__":
    main()