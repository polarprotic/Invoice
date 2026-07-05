"""
Invoice Extraction Pipeline -- Puter (free, OpenAI-compatible API), with dynamic fields
------------------------------------------------------------------------------
Same schema logic, caching, and Excel outputs as the Ollama/Gemini pipelines --
only the model backend differs. This one talks to Puter's OpenAI-compatible
REST API (https://api.puter.com/puterai/openai/v1/) using the standard
`openai` Python package, pointed at a custom base_url. No Node.js, no
browser, no Puter.js JS library required -- pure Python.

  - CORE fields (document_type, date, payer, payee, amount, reference_no,
    bank, account_no, notes) are always extracted and always appear as the
    first columns.
  - Anything else the model finds (discount, tax, GST number, due date, PO
    number, freight, TDS, cheque validity date, etc.) goes into an
    "extra_fields" object, unioned across the whole batch into extra
    columns (capped at MAX_EXTRA_COLUMNS, overflow folds into notes).
  - RESULTS ARE CACHED per file (raw_extraction.json) -- re-running the
    script never re-calls the API for a file that's already been
    extracted. Only the Excel layout gets rebuilt every run.
  - A combined "All_Invoices_Combined.xlsx" master sheet with every record.
  - A "All_Invoices_Visual_Report.xlsx" with a thumbnail of each source
    image next to a table of just that invoice's own fields.

MODEL FALLBACK:
  PUTER_MODELS is tried in order -- if a model fails all its retries on a
  given file (rate limits, transient errors, refusals, etc.) the script
  automatically tries the next model in the list before giving up on that
  file. Since Puter doesn't publish a documented daily quota the way
  Gemini's free tier does, there's no global "stop the whole run" quota
  exhaustion here -- a file that fails on every model in the list is just
  logged as failed and picked back up automatically next run (no cache
  gets written for it).

SETUP:
    pip install -r requirements_puter.txt
    (or manually: pip install openai pymupdf pillow pandas openpyxl)

    Get a free Puter auth token:
      1. Sign up at https://puter.com
      2. Go to https://puter.com/dashboard#account
      3. Find "API token" -> click "Create token" -> copy it
      4. export PUTER_AUTH_TOKEN="your-token-here"
         (Windows PowerShell: $env:PUTER_AUTH_TOKEN = "your-token-here")

USAGE:
    python invoice_pipeline_puter.py

    (reads from ./input_img, writes to ./invoice_output/<filename>/)
"""

import os
import io
import re
import json
import time
import glob
import base64
import shutil
import tempfile

from openai import OpenAI
import openai as openai_errors
from PIL import Image
import fitz  # PyMuPDF
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- CONFIG ----------
INPUT_DIR = "input_img"
OUTPUT_ROOT = "invoice_output"

PUTER_BASE_URL = "https://api.puter.com/puterai/openai/v1/"

# Tried in order for each file. First is primary; if it fails all its
# retries the script falls back to the next one before giving up on the
# file entirely. Mix vision-capable models from different underlying
# providers so a bad day for one doesn't stall the whole batch.
PUTER_MODELS = [
    "qwen/qwen3-vl-30b-a3b",
    "gemini-3.1-flash-lite",
    "gpt-5.4-nano",
]

SLEEP_BETWEEN_CALLS = 2
MAX_RETRIES = 3
REQUEST_TIMEOUT = 180  # seconds

# Gemini/GPT/Qwen all cost tokens proportional to image resolution -- keep
# this modest even though Puter doesn't publish a hard daily quota, since
# large images are still slower per-request and more likely to trip
# whatever throttling exists under the hood.
MAX_IMAGE_DIMENSION = 1600
PDF_RENDER_DPI = 150

# Cap on how many distinct extra-field columns a single file's sheet can
# grow to. Guards against a garbled model response spraying dozens of junk
# keys into the header row. Anything beyond the cap collapses into "notes".
MAX_EXTRA_COLUMNS = 25

# Keys (after normalize_key) that should be written as real numbers rather
# than text, so Excel SUM()/number formatting works on them instead of
# silently treating them as blank.
NUMERIC_KEYS = {"subtotal", "tax", "discount", "tds", "freight",
                 "balance_remaining"}

# Visual report layout: thumbnail width in pixels, blank rows between one
# invoice's block (image + table) and the next, and the row-height estimate
# used to reserve enough rows under an image before the table starts.
VISUAL_THUMB_MAX_DIM = 220
VISUAL_GAP_ROWS = 4
EXCEL_DEFAULT_ROW_HEIGHT_PX = 20  # ~15pt default row height in pixels
# -----------------------------

CORE_COLUMNS = ["source_file", "document_type", "date", "payer", "payee",
                "amount", "reference_no", "bank", "account_no", "notes"]

# Known synonyms get normalized to one canonical key so "Discount %" and
# "discount_percent" don't become two separate columns. Match is
# case/whitespace-insensitive against the lowercased, underscored key the
# model returns.
FIELD_ALIASES = {
    "discount": "discount",
    "discount_amount": "discount",
    "discount_percent": "discount",
    "gst": "gst_no",
    "gst_no": "gst_no",
    "gst_number": "gst_no",
    "gstin": "gst_no",
    "tax": "tax",
    "tax_amount": "tax",
    "tds": "tds",
    "due_date": "due_date",
    "payment_due_date": "due_date",
    "po_number": "po_number",
    "purchase_order_no": "po_number",
    "freight": "freight",
    "shipping_charges": "freight",
    "cheque_validity_date": "cheque_validity_date",
    "ifsc_code": "ifsc_code",
    "ifsc": "ifsc_code",
}


def normalize_key(k):
    k = re.sub(r"[^a-z0-9]+", "_", k.strip().lower()).strip("_")
    return FIELD_ALIASES.get(k, k)


BASE_PROMPT = """Extract structured data from this Indian bank/advertising payment
document. It may contain a MIX of typed/printed text and handwritten text in
the same image (e.g. a printed form filled in by hand, or a receipt with a
handwritten section next to a printed letterhead). Read both carefully,
using the actual image as ground truth.

Return a JSON array. Each element = ONE distinct financial transaction or
record found anywhere in the image (a single image may show multiple cheques,
receipts, or line-items -- list each one separately, do not merge them).

For each record extract these CORE fields (always include all of them, use ""
if missing/illegible):
- document_type: e.g. "Cheque", "Deposit Slip", "NEFT Advice", "Payment Voucher",
  "Release Order", "Receipt", "Tax Invoice", "Invoice"
- date: exactly as written (date only -- put any separate time value in extra_fields, see below)
- payer: who is paying / sending money / the client being billed
- payee: who receives the money / the business issuing the document
- amount: the FINAL/net payable or total amount, number only, no currency
  symbols, no commas, no units (e.g. "540" not "USD540" or "540.00 USD")
- reference_no: cheque no. / UTR / transaction ID / receipt no. / invoice no. / bill no. --
  whichever is present (combine if multiple, comma separated)
- bank: bank name involved (leave "" if not a banking document)
- account_no: as shown, including masked digits if partially hidden
- notes: only for genuinely unstructured remarks that don't fit as a labeled
  field anywhere else (e.g. a free-text comment) -- do NOT use notes as a
  dumping ground for labeled fields, those belong in extra_fields below

CRITICAL -- BE EXHAUSTIVE: this document may be one of thousands of
different formats (invoices, receipts, cheques, vouchers, advices) and each
one can carry different labeled information. Do not limit yourself to a
fixed list. Scan the ENTIRE document top to bottom and capture EVERY
distinct labeled piece of information you can see as a key in
"extra_fields", even if it seems minor or you're unsure it matters.
Concretely, always check for (include only if actually present on the
document):
- time (if a timestamp with a time component appears, separate from date)
- currency / currency_unit (e.g. "USD", "INR") if the amount has one
- subtotal, tax, discount, gst_no, tds, freight, balance_remaining,
  amount_in_words
- invoice_no (if distinct from reference_no), po_number, due_date,
  payment_method, rep / prepared_by / issued_by, client_address,
  company_address, company_phone, company_email, ifsc_code,
  branch, cheque_validity_date
- ANY other labeled field visible on the document that isn't covered above
  -- invent a short snake_case key for it rather than skipping it or
  folding it into notes.

Use short snake_case keys and plain values -- strip currency symbols/commas
from numeric-looking values (put the currency itself in a separate
"currency" key if shown). Only include keys for data actually present in
THIS document -- do not invent data, and leave "extra_fields" as {} only if
there is truly nothing beyond the core fields.

Return ONLY the JSON array -- no markdown fences, no explanation, no
commentary before or after the array. Example shape:
[{"document_type": "...", "date": "...", "payer": "...", "payee": "...",
  "amount": "540", "reference_no": "...", "bank": "...", "account_no": "...",
  "notes": "...", "extra_fields": {"time": "11:10", "currency": "USD",
  "subtotal": "540", "tax": "0", "balance_remaining": "0",
  "payment_method": "Bank Transfer", "rep": "J. Nolan"}}]"""


def load_pages(path):
    ext = path.lower().rsplit(".", 1)[-1]
    if ext == "pdf":
        doc = fitz.open(path)
        return [Image.open(io.BytesIO(page.get_pixmap(dpi=PDF_RENDER_DPI).tobytes("png"))) for page in doc]
    return [Image.open(path).convert("RGB")]


def resize_for_model(img, max_dim=MAX_IMAGE_DIMENSION):
    w, h = img.size
    longest = max(w, h)
    if longest <= max_dim:
        return img
    scale = max_dim / longest
    new_size = (int(w * scale), int(h * scale))
    return img.resize(new_size, Image.LANCZOS)


def pil_to_data_uri(img):
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=90)
    b64 = base64.b64encode(buf.getvalue()).decode("utf-8")
    return f"data:image/jpeg;base64,{b64}"


def extract_json(text):
    text = text.strip()

    # Some models wrap the answer in a reasoning block or code fence --
    # strip/locate the JSON robustly instead of assuming clean output.
    text = re.sub(r"<think>.*?</think>", "", text, flags=re.DOTALL).strip()

    try:
        return json.loads(text)
    except Exception:
        pass

    fenced = re.search(r"```(?:json)?\s*(.*?)```", text, re.DOTALL)
    if fenced:
        try:
            return json.loads(fenced.group(1).strip())
        except Exception:
            pass

    start = text.find("[")
    if start != -1:
        depth = 0
        for i in range(start, len(text)):
            if text[i] == "[":
                depth += 1
            elif text[i] == "]":
                depth -= 1
                if depth == 0:
                    candidate = text[start:i + 1]
                    try:
                        return json.loads(candidate)
                    except Exception:
                        break

    return None


def check_puter_auth(client):
    try:
        client.chat.completions.create(
            model=PUTER_MODELS[0],
            messages=[{"role": "user", "content": "Reply with just: ok"}],
            timeout=30,
        )
        return True
    except openai_errors.AuthenticationError:
        print("ERROR: Puter auth token was rejected (401 Unauthorized).")
        print("  - Check PUTER_AUTH_TOKEN is set correctly.")
        print("  - Get a fresh token at https://puter.com/dashboard#account")
        return False
    except Exception as e:
        # Anything else (rate limit, model hiccup, etc.) isn't an auth
        # problem -- don't block startup over a transient error here.
        print(f"    [warmup call had an issue, continuing anyway: {e}]")
        return True


def extract(path, images, client, debug_dir=None):
    images_data_uri = [pil_to_data_uri(resize_for_model(img)) for img in images]
    content = [{"type": "text", "text": BASE_PROMPT}]
    for uri in images_data_uri:
        content.append({"type": "image_url", "image_url": {"url": uri}})

    for model_name in PUTER_MODELS:
        for attempt in range(1, MAX_RETRIES + 1):
            try:
                resp = client.chat.completions.create(
                    model=model_name,
                    messages=[{"role": "user", "content": content}],
                    temperature=0.1,
                    timeout=REQUEST_TIMEOUT,
                )
                raw = resp.choices[0].message.content or ""
                parsed = extract_json(raw)
                if parsed is not None:
                    return parsed
                print(f"    [parse failed, retry {attempt}/{MAX_RETRIES}] "
                      f"{model_name} didn't return valid JSON")
                if debug_dir:
                    os.makedirs(debug_dir, exist_ok=True)
                    dump_path = os.path.join(
                        debug_dir, f"raw_response_{model_name.replace('/', '_')}_attempt{attempt}.txt")
                    with open(dump_path, "w", encoding="utf-8") as fh:
                        fh.write(raw)
                    print(f"        raw response saved to {dump_path}")

            except openai_errors.AuthenticationError as e:
                # Bad/expired token -- retrying or falling back to another
                # model won't help, this needs a human to fix the token.
                print(f"    [AUTH ERROR] Puter token rejected: {e}")
                print(f"    [FAILED] {os.path.basename(path)} -- skipped, needs manual entry")
                return []

            except openai_errors.RateLimitError as e:
                print(f"    [rate limited on {model_name}, backing off] "
                      f"attempt {attempt}/{MAX_RETRIES}: {e}")
                time.sleep(15 * attempt)
                continue

            except (openai_errors.APIConnectionError, openai_errors.APITimeoutError) as e:
                print(f"    [connection/timeout on {model_name}, retry "
                      f"{attempt}/{MAX_RETRIES}] {e}")
                time.sleep(10 * attempt)
                continue

            except Exception as e:
                print(f"    [retry {attempt}/{MAX_RETRIES} on {model_name}] {e}")
                time.sleep(5 * attempt)

        remaining = [m for m in PUTER_MODELS if m != model_name]
        idx = PUTER_MODELS.index(model_name)
        if idx + 1 < len(PUTER_MODELS):
            print(f"    [falling back from {model_name} to {PUTER_MODELS[idx + 1]}]")

    print(f"    [FAILED] {os.path.basename(path)} -- skipped, needs manual entry")
    return []


def sanitize_cell_value(v):
    """Excel cells can only hold plain scalars (str/int/float/bool/None) --
    not lists or dicts. Models occasionally return a list for a field that
    should have been a string (e.g. an itemized list of products as the
    value for one extra_fields key). Coerce anything non-scalar into a
    readable string instead of crashing the Excel writer downstream."""
    if isinstance(v, (list, tuple)):
        return "; ".join(sanitize_cell_value(x) for x in v)
    if isinstance(v, dict):
        return "; ".join(f"{k}: {sanitize_cell_value(val)}" for k, val in v.items())
    return v


def flatten_records(results, source_file):
    """Split each raw model record into its core dict + normalized
    extra_fields dict. No capping here -- this file's fields get unioned
    with every other file's fields first, and the cap (if any) is applied
    once globally in main() so column layout stays consistent batch-wide."""
    core_rows = []

    for r in results:
        core = {col: sanitize_cell_value(r.get(col, "")) for col in CORE_COLUMNS}
        core["source_file"] = source_file

        raw_extra = r.get("extra_fields") or {}
        norm_extra = {}
        if isinstance(raw_extra, dict):
            for k, v in raw_extra.items():
                nk = normalize_key(str(k))
                if nk:
                    norm_extra[nk] = sanitize_cell_value(v)

        core["_extra"] = norm_extra
        core_rows.append(core)

    return core_rows


def build_global_extra_columns(all_core_rows):
    """Union extra_fields keys across EVERY row in the whole batch, in
    first-seen order. Cap applied once, globally -- fields beyond the cap
    fold into that row's notes instead of getting a column."""
    freq = {}
    order = []
    for core in all_core_rows:
        for k in core.get("_extra", {}):
            if k not in freq:
                freq[k] = 0
                order.append(k)
            freq[k] += 1

    ranked = sorted(order, key=lambda k: (-freq[k], order.index(k)))
    kept = ranked[:MAX_EXTRA_COLUMNS]
    dropped = set(ranked[MAX_EXTRA_COLUMNS:])

    if dropped:
        for core in all_core_rows:
            extra = core.get("_extra", {})
            overflow = {k: v for k, v in extra.items() if k in dropped}
            if overflow:
                extra_str = "; ".join(f"{k}: {v}" for k, v in overflow.items())
                core["notes"] = (core["notes"] + "; " if core["notes"] else "") + extra_str

    return [k for k in order if k in set(kept)]


def style_workbook(path, amt_col_letter, n_core_cols, n_extra_cols):
    wb = load_workbook(path)
    ws = wb.active
    n_cols = ws.max_column

    header_fill = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
    extra_header_fill = PatternFill("solid", start_color="2E7D32", end_color="2E7D32")
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, n_cols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = extra_header_fill if col > n_core_cols else header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    for r in range(2, ws.max_row + 1):
        for col in range(1, n_cols + 1):
            c = ws.cell(row=r, column=col)
            c.font = Font(name="Arial", size=10)
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if col == amt_col_letter:
                c.number_format = "#,##0.00"
                c.alignment = Alignment(horizontal="right", vertical="center")

    core_widths = [30, 22, 14, 26, 26, 14, 36, 20, 18, 36]
    for i, w in enumerate(core_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for i in range(n_core_cols + 1, n_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18
    ws.freeze_panes = "A2"

    if ws.max_row > 1:
        last_row = ws.max_row
        total_row = last_row + 2
        col_letter = get_column_letter(amt_col_letter)
        label_col = max(1, amt_col_letter - 1)
        ws.cell(row=total_row, column=label_col, value="TOTAL").font = Font(name="Arial", bold=True)
        tc = ws.cell(row=total_row, column=amt_col_letter,
                     value=f"=SUM({col_letter}2:{col_letter}{last_row})")
        tc.number_format = "#,##0.00"
        tc.font = Font(name="Arial", bold=True)

    wb.save(path)


def write_visual_report(files, per_file_rows, out_path):
    """Combined workbook where each invoice gets its own block: a thumbnail
    of the source image, then a table built from ONLY that invoice's own
    fields (core fields that had a value + whatever extra_fields it had)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    header_fill = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    filename_font = Font(name="Arial", bold=True, size=12)
    cell_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row_cursor = 1
    max_cols_used = 1
    tmp_dir = tempfile.mkdtemp(prefix="invoice_thumbs_")

    try:
        for f in files:
            base_name = os.path.splitext(os.path.basename(f))[0]
            rows = per_file_rows.get(base_name)
            if not rows:
                continue

            block_start_row = row_cursor

            img_rows_used = 6
            try:
                pages = load_pages(f)
                thumb = pages[0].copy()
                thumb.thumbnail((VISUAL_THUMB_MAX_DIM, VISUAL_THUMB_MAX_DIM), Image.LANCZOS)
                thumb_path = os.path.join(tmp_dir, f"{base_name}.png")
                thumb.save(thumb_path)
                xl_img = XLImage(thumb_path)
                ws.add_image(xl_img, f"A{row_cursor}")
                img_rows_used = max(6, (xl_img.height // EXCEL_DEFAULT_ROW_HEIGHT_PX) + 2)
            except Exception as e:
                ws.cell(row=row_cursor, column=1, value=f"[thumbnail unavailable: {e}]").font = cell_font

            table_row = block_start_row + img_rows_used

            ws.cell(row=table_row, column=1, value=os.path.basename(f)).font = filename_font
            table_row += 1

            for rec in rows:
                cols = [c for c in CORE_COLUMNS if c != "source_file" and rec.get(c, "")]
                extra = rec.get("_extra", {})
                extra_cols = [k for k, v in extra.items() if v not in ("", None)]
                cols += extra_cols
                if not cols:
                    continue
                max_cols_used = max(max_cols_used, len(cols))

                for j, c in enumerate(cols, start=1):
                    hc = ws.cell(row=table_row, column=j, value=c.replace("_", " ").title())
                    hc.font = header_font
                    hc.fill = header_fill
                    hc.border = border
                    hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                table_row += 1

                for j, c in enumerate(cols, start=1):
                    val = rec.get(c, "") if c in CORE_COLUMNS else extra.get(c, "")
                    vc = ws.cell(row=table_row, column=j, value=val)
                    vc.font = cell_font
                    vc.border = border
                    vc.alignment = Alignment(vertical="center", wrap_text=True)
                table_row += 1

                table_row += 1

            row_cursor = max(table_row, block_start_row + img_rows_used + 2) + VISUAL_GAP_ROWS

        for i in range(1, max_cols_used + 1):
            ws.column_dimensions[get_column_letter(i)].width = 20

        wb.save(out_path)
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


def write_master_excel(all_core_rows, extra_columns, out_path):
    write_excel(all_core_rows, extra_columns, out_path)


def write_excel(core_rows, extra_columns, out_path):
    all_columns = CORE_COLUMNS + extra_columns
    rows = []
    for core in core_rows:
        row = {col: core.get(col, "") for col in CORE_COLUMNS}
        extra = core.get("_extra", {})
        for ec in extra_columns:
            row[ec] = extra.get(ec, "")
        rows.append(row)

    df = pd.DataFrame(rows, columns=all_columns)

    def to_number(v):
        if v in ("", None):
            return None
        try:
            return float(str(v).replace(",", "").strip())
        except ValueError:
            return v

    df["amount"] = df["amount"].apply(to_number)
    for col in extra_columns:
        if col in NUMERIC_KEYS:
            df[col] = df[col].apply(to_number)

    try:
        df.to_excel(out_path, index=False)
        amt_col_idx = CORE_COLUMNS.index("amount") + 1
        style_workbook(out_path, amt_col_idx, len(CORE_COLUMNS), len(extra_columns))
    except PermissionError:
        raise PermissionError(
            f"'{out_path}' is open in Excel (or similar) -- close it and re-run "
            f"the script later to pick this file up."
        )


def main():
    auth_token = os.environ.get("PUTER_AUTH_TOKEN")
    if not auth_token:
        print("ERROR: PUTER_AUTH_TOKEN is not set in this terminal session.")
        print("  1. Sign up at https://puter.com")
        print("  2. Go to https://puter.com/dashboard#account")
        print("  3. Find 'API token' -> click 'Create token' -> copy it")
        print("  Windows PowerShell (current session only): $env:PUTER_AUTH_TOKEN = \"your-token-here\"")
        print("  Mac/Linux: export PUTER_AUTH_TOKEN=\"your-token-here\"")
        return

    client = OpenAI(base_url=PUTER_BASE_URL, api_key=auth_token)

    print("=" * 60)
    print("CONFIG IN USE (check this matches what you expect):")
    print(f"  PUTER_BASE_URL      = {PUTER_BASE_URL}")
    print(f"  PUTER_MODELS        = {' -> '.join(PUTER_MODELS)}")
    print(f"  MAX_IMAGE_DIMENSION = {MAX_IMAGE_DIMENSION}px")
    print(f"  MAX_EXTRA_COLUMNS   = {MAX_EXTRA_COLUMNS}")
    print("=" * 60 + "\n")

    print("Checking Puter auth token...")
    if not check_puter_auth(client):
        return
    print("Auth OK.\n")

    files = sorted(
        f for f in glob.glob(os.path.join(INPUT_DIR, "*"))
        if f.lower().endswith((".jpg", ".jpeg", ".png", ".pdf"))
    )
    if not files:
        print(f"No .jpg/.jpeg/.png/.pdf files found in '{INPUT_DIR}/'")
        return

    os.makedirs(OUTPUT_ROOT, exist_ok=True)

    def raw_cache_path(f):
        base_name = os.path.splitext(os.path.basename(f))[0]
        return os.path.join(OUTPUT_ROOT, base_name, "raw_extraction.json")

    def already_extracted(f):
        return os.path.exists(raw_cache_path(f))

    pending = [f for f in files if not already_extracted(f)]
    cached_count = len(files) - len(pending)
    if cached_count:
        print(f"{cached_count} file(s) already extracted (using cached results).")
    if pending:
        print(f"{len(pending)} file(s) to extract via Puter "
              f"({' -> '.join(PUTER_MODELS)}).\n")
    else:
        print(f"Nothing left to extract -- all {len(files)} file(s) already cached.")

    failed = []

    # ---- PASS 1: extract (or load from cache) every file's raw records ----
    for i, f in enumerate(pending, 1):
        base_name = os.path.splitext(os.path.basename(f))[0]
        print(f"[{i}/{len(pending)}] Extracting {os.path.basename(f)}")

        folder = os.path.join(OUTPUT_ROOT, base_name)
        os.makedirs(folder, exist_ok=True)

        try:
            shutil.copy2(f, os.path.join(folder, os.path.basename(f)))
            images = load_pages(f)
            results = extract(f, images, client, debug_dir=folder)
            tmp_path = raw_cache_path(f) + ".tmp"
            with open(tmp_path, "w", encoding="utf-8") as fh:
                json.dump(results, fh, ensure_ascii=False, indent=2)
            os.replace(tmp_path, raw_cache_path(f))
            print(f"    -> {len(results)} record(s) extracted")
        except Exception as e:
            print(f"    [SKIPPED - error] {os.path.basename(f)}: {e}")
            failed.append(os.path.basename(f))

        time.sleep(SLEEP_BETWEEN_CALLS)

    # ---- PASS 2: rebuild every sheet from cache with a consistent, ----
    # ---- batch-wide column layout (covers cached files from earlier runs too) ----
    print("\nBuilding spreadsheets with a consistent column layout across the batch...")

    per_file_rows = {}
    for f in files:
        cache_path = raw_cache_path(f)
        if not os.path.exists(cache_path):
            continue
        base_name = os.path.splitext(os.path.basename(f))[0]
        try:
            with open(cache_path, "r", encoding="utf-8") as fh:
                results = json.load(fh)
        except json.JSONDecodeError:
            print(f"    [BAD CACHE] {base_name}: raw_extraction.json is corrupted, "
                  f"deleting it -- re-run the script to re-extract this file.")
            os.remove(cache_path)
            failed.append(os.path.basename(f))
            continue
        per_file_rows[base_name] = flatten_records(results, os.path.basename(f))

    all_core_rows = [row for rows in per_file_rows.values() for row in rows]
    if not all_core_rows:
        print("No successfully extracted records to write.")
    else:
        extra_columns = build_global_extra_columns(all_core_rows)
        print(f"Global extra columns ({len(extra_columns)}): {', '.join(extra_columns) if extra_columns else '(none)'}")

        for base_name, rows in per_file_rows.items():
            out_xlsx = os.path.join(OUTPUT_ROOT, base_name, "Extracted_Invoice.xlsx")
            try:
                write_excel(rows, extra_columns, out_xlsx)
            except PermissionError as e:
                print(f"    [SKIPPED WRITE] {base_name}: {e}")

        master_path = os.path.join(OUTPUT_ROOT, "All_Invoices_Combined.xlsx")
        try:
            write_master_excel(all_core_rows, extra_columns, master_path)
            print(f"\nMaster combined workbook: {master_path} ({len(all_core_rows)} record(s) across {len(per_file_rows)} file(s))")
        except PermissionError as e:
            print(f"    [SKIPPED WRITE] master workbook: {e}")

        visual_path = os.path.join(OUTPUT_ROOT, "All_Invoices_Visual_Report.xlsx")
        try:
            write_visual_report(files, per_file_rows, visual_path)
            print(f"Visual report (thumbnail + per-invoice table): {visual_path}")
        except PermissionError:
            print(f"    [SKIPPED WRITE] visual report: '{visual_path}' is open elsewhere -- close it and re-run.")
        except Exception as e:
            print(f"    [SKIPPED WRITE] visual report failed: {e}")

    print(f"\nDone. Per-file results in '{OUTPUT_ROOT}/<filename>/Extracted_Invoice.xlsx'.")
    if failed:
        print(f"{len(failed)} file(s) failed extraction (re-run the script "
              f"to retry them automatically): {', '.join(failed)}")


if __name__ == "__main__":
    main()